import pandas as pd
from Accessdb import AccessHelper
import os
import random
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import warnings

# ==========================================
# 0. 全域檔案、路徑與門檻設定 (置於最前端便於修改)
# ==========================================
# 📌 外部備援分類表路徑設定 (當資料庫 Courses 表找不到時自動啟用)
BACKUP_CLASS_FILE = r'D:\113年後資料\系辦辦公相關\IEET認證\python程式\PythonIEET\PythonIEET\input_files\課程分類表\課程分類表_20260610.xlsx'

# 📌 畢業學分檢核基準數設定
MIN_GRAD_CREDITS_UNDERGRAD = 128  # 大學部最低畢業總學分基準 (128學分)
MIN_GRAD_CREDITS_GRAD = 24        # 碩士班最低畢業總學分基準

# 📌 輸出目錄設定
OUTPUT_BASE_DIR = 'output_files'
OUTPUT_SUB_DIR = '畢業生成績分析'


# ==========================================
# 1. 環境設定
# ==========================================
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)

db = AccessHelper()
today_str = datetime.today().strftime('%Y%m%d')

# 自動建立輸出路徑
OUTPUT_DIR_PATH = os.path.join(OUTPUT_BASE_DIR, OUTPUT_SUB_DIR)
if not os.path.exists(OUTPUT_DIR_PATH):
    os.makedirs(OUTPUT_DIR_PATH)
    print(f"建立目錄: {OUTPUT_DIR_PATH}")


# ==========================================
# 2. 輔助函數
# ==========================================
def to_str(val):
    if pd.isna(val): return ""
    s = str(val).strip()
    return s[:-2] if s.endswith('.0') else s

def to_int(val, default=0):
    try: return int(float(str(val)))
    except: return default

def find_col(df, keywords):
    for col in df.columns:
        if any(k in col.lower() for k in keywords): return col
    return None

def check_bool(val):
    return str(val).strip().lower() in ['1', 'true', 'yes', 'v', '1.0', '1']

def get_course_code_column(df):
    candidates = ['course_code', '課號', 'code', 'c_code']
    for col in df.columns:
        if col.lower() in candidates: return col
    return None

def number_to_chinese_grade(n):
    try:
        n = int(n)
        mapping = {1: '一', 2: '二', 3: '三', 4: '四', 5: '五', 6: '六', 7: '七', 8: '八'}
        return mapping.get(n, str(n))
    except:
        return '預' if str(n).lower() in ['pre', '0'] else str(n)

def number_to_chinese_sem(n):
    try:
        n = int(n)
        mapping = {1: '上', 2: '下', 3: '補'}
        return mapping.get(n, str(n))
    except:
        return str(n)

def format_semester_key(key):
    try:
        g_str, s_str = key.replace(')', '').split('(')
        c_grade = number_to_chinese_grade(g_str)
        c_sem = number_to_chinese_sem(s_str)
        return f"{c_grade}({c_sem})"
    except:
        return key

# 百分比計算輔助函數
def calc_pct(val, total):
    if total == 0: return "0%"
    return f"{val/total:.1%}"

# 判斷學生開始修課年級與學期的函數
def determine_start_grade(student_rec):
    for g in range(1, 8):
        for s in range(1, 3):
            col_name = f"Y{g}S{s}_Cred"
            if col_name in student_rec:
                val = student_rec[col_name]
                try:
                    if float(val) > 0: return g, s
                except: pass
    return 1, 1 

# 💡 [全新優化] 檔名防覆蓋輔助函數：若當天已有名稱相同的檔案，自動加上序號 (_1, _2...)
def get_unique_filepath(dir_path, filename):
    base, ext = os.path.splitext(filename)
    full_path = os.path.join(dir_path, filename)
    counter = 1
    while os.path.exists(full_path):
        new_filename = f"{base}_{counter}{ext}"
        full_path = os.path.join(dir_path, new_filename)
        counter += 1
    return full_path


# ==========================================
# 3. 資料載入 (排名主表)
# ==========================================
print("讀取隨機抽樣排名資料表...")
df_rank_u_all = pd.read_sql("SELECT * FROM GradRankU", db.conn)
df_rank_g_all = pd.read_sql("SELECT * FROM GradRankG", db.conn)

id_col_u = 'StudentID'
ay_col_u = 'AcademicYear'
class_col = 'Class'
rank_col_u = 'Rank'
grade_col_u = 'Grade'
total_col_u = find_col(df_rank_u_all, ['TotalStudents', '總人數'])

id_col_g = 'StudentID'
ay_col_g = 'AcademicYear'
rank_col_g = 'Rank'
grade_col_g = 'Grade'
total_col_g = find_col(df_rank_g_all, ['TotalStudents', '總人數'])

for df in [df_rank_u_all, df_rank_g_all]:
    for col in [id_col_u, ay_col_u, id_col_g, ay_col_g]: 
        if col in df.columns: df[col] = df[col].apply(to_str)


# ==========================================
# 4. 分層篩選隨機抽樣函式 (新增轉學生過濾關卡)
# ==========================================
def get_stratified_samples(df, ay_col, class_col, rank_col, total_col, id_col, grade_col, is_undergrad=True):
    target_years = [str(y) for y in range(109, 115)]
    df_pool = df[df[ay_col].isin(target_years)].copy()
    
    if is_undergrad and class_col in df.columns:
        df_pool = df_pool[df_pool[class_col] == '全']
        
    # 💡 [全新優化] 排除轉學生：檢查 EntryChannel 欄位，若為 '轉學考' 則直接從抽樣母體中排除，避免抽到轉學生
    if 'EntryChannel' in df_pool.columns:
        df_pool = df_pool[df_pool['EntryChannel'].astype(str).str.strip() != '轉學考']
        
    if not total_col:
        id_c = id_col_u if is_undergrad else id_col_g
        counts = df_pool.groupby(ay_col)[id_c].count().reset_index()
        counts.rename(columns={id_c: 'calc_total'}, inplace=True)
        df_pool = df_pool.merge(counts, on=ay_col, how='left')
        total_col = 'calc_total'
        
    df_pool[rank_col] = pd.to_numeric(df_pool[rank_col], errors='coerce')
    df_pool[total_col] = pd.to_numeric(df_pool[total_col], errors='coerce')
    df_pool = df_pool[df_pool[total_col] > 0]

    result = {}
    sampled_years = []
    years = sorted(df_pool[ay_col].unique())
    
    for year in years:
        df_year = df_pool[df_pool[ay_col] == year]
        if df_year.empty: continue
            
        total_students = df_year.iloc[0][total_col]
        t1 = total_students * 0.33
        t2 = total_students * 0.66
        
        group_top = df_year[df_year[rank_col] <= t1]
        group_mid = df_year[(df_year[rank_col] > t1) & (df_year[rank_col] <= t2)]
        group_bot = df_year[df_year[rank_col] > t2]
        
        selected_records = []
        
        def pick(group, label):
            n = min(2, len(group))
            if n > 0:
                picks = group.sample(n=n)
                for _, row in picks.iterrows():
                    rec = row.to_dict()
                    rec['sid'] = row[id_col]
                    rec['rank'] = row[rank_col]
                    rec['total'] = row[total_col]
                    rec['grad_year'] = row[ay_col]
                    rec['grad_grade'] = row[grade_col] if grade_col in row else None
                    rec['group_label'] = label
                    selected_records.append(rec)

        pick(group_top, "前段")
        pick(group_mid, "中段")
        pick(group_bot, "後段")
        
        if selected_records:
            selected_records.sort(key=lambda x: x['rank'])
            result[year] = selected_records
            sampled_years.append(year)
            
    return result, sampled_years

print("正在進行畢業生分層抽樣 (已排除轉學考生)...")
u_data, u_years = get_stratified_samples(df_rank_u_all, ay_col_u, class_col, rank_col_u, total_col_u, id_col_u, grade_col_u, True)
g_data, g_years = get_stratified_samples(df_rank_g_all, ay_col_g, None, rank_col_g, total_col_g, id_col_g, grade_col_g, False)

all_ids = []
for d in [u_data, g_data]:
    for yr in d:
        all_ids.extend([x['sid'] for x in d[yr]])

if not all_ids:
    print("無符合抽樣之學生，程式結束。")
    exit()


# ==========================================
# 5. 建立雙軌制課程分類對應表 (資料庫表 + 外部備援表)
# ==========================================
print("從資料庫載入電機系開課主表 (Courses)...")
sql_courses = """
    SELECT academic_year, semester, course_code, course_name, is_required, credits, instructor, 
           is_math, is_science, is_eng_prof, is_general 
    FROM Courses
"""
df_db_courses = pd.read_sql(sql_courses, db.conn)

db_code_map = {}
db_name_map = {}
for _, row in df_db_courses.iterrows():
    cats = {
        'math': bool(row['is_math']),
        'sci': bool(row['is_science']),
        'eng': bool(row['is_eng_prof']),
        'gen': bool(row['is_general'])
    }
    if pd.notna(row['course_code']):
        db_code_map[str(row['course_code']).strip()] = cats
    if pd.notna(row['course_name']):
        db_name_map[str(row['course_name']).strip()] = cats

ext_code_map = {}
ext_name_map = {}
if os.path.exists(BACKUP_CLASS_FILE):
    print(f"✅ 成功尋獲並載入外部備援分類表: {BACKUP_CLASS_FILE}")
    try:
        df_ext_map = pd.read_excel(BACKUP_CLASS_FILE)
        df_ext_map.columns = [c.strip() for c in df_ext_map.columns]
        
        col_c_code = next((c for c in df_ext_map.columns if '課號' in c or 'code' in c), None)
        col_c_name = next((c for c in df_ext_map.columns if '名稱' in c or 'name' in c), None)
        col_m = next((c for c in df_ext_map.columns if '數學' in c or 'math' in c), None)
        col_s = next((c for c in df_ext_map.columns if '科學' in c or 'sci' in c), None)
        col_e = next((c for c in df_ext_map.columns if '工程' in c or 'eng' in c), None)
        col_g = next((c for c in df_ext_map.columns if '通識' in c or 'gen' in c), None)
        
        for _, row in df_ext_map.iterrows():
            cats_ext = {
                'math': check_bool(row.get(col_m, 0)),
                'sci': check_bool(row.get(col_s, 0)),
                'eng': check_bool(row.get(col_e, 0)),
                'gen': check_bool(row.get(col_g, 0))
            }
            if col_c_code and pd.notna(row[col_c_code]):
                ext_code_map[str(row[col_c_code]).strip()] = cats_ext
            if col_c_name and pd.notna(row[col_c_name]):
                ext_name_map[str(row[col_c_name]).strip()] = cats_ext
        print(f"➔ 備援分類表資料記憶化完成 (共取得 {max(len(ext_code_map), len(ext_name_map))} 筆獨立對應組合)")
    except Exception as e:
        print(f"⚠️ 讀取外部備援表失敗，分析將全數依賴資料庫: {e}")
else:
    print(f"⚠️ 未尋獲外部備援檔 {BACKUP_CLASS_FILE}，將全數採用資料庫進行比對。")

df_scores_raw = pd.read_sql("SELECT * FROM STscore", db.conn)
sc_id_col = find_col(df_scores_raw, ['學號', 'StudentID'])
sc_ay_col = find_col(df_scores_raw, ['學年度', 'Year'])
sc_sem_col = find_col(df_scores_raw, ['學期', 'Semester'])
sc_code_col = find_col(df_scores_raw, ['課號', 'Code'])
sc_name_col = find_col(df_scores_raw, ['名稱', 'Name', 'Title'])
sc_credit_col = find_col(df_scores_raw, ['學分', 'Credit'])
sc_score_col = find_col(df_scores_raw, ['成績', 'Score']) 

df_scores_raw[sc_id_col] = df_scores_raw[sc_id_col].apply(to_str)
df_scores_raw[sc_ay_col] = df_scores_raw[sc_ay_col].apply(to_str)
df_scores_raw[sc_sem_col] = df_scores_raw[sc_sem_col].apply(to_str)

df_scores = df_scores_raw[df_scores_raw[sc_id_col].isin(all_ids)].copy()
missing_courses = set()


# ==========================================
# 6. 雙軌制分類提領與成績審查累加邏輯
# ==========================================
def get_course_category_dual(c_name, c_code):
    c_name = str(c_name).strip()
    c_code = str(c_code).strip()
    
    if c_code in db_code_map: return db_code_map[c_code]
    if c_name in db_name_map: return db_name_map[c_name]
        
    if c_code in ext_code_map: return ext_code_map[c_code]
    if c_name in ext_name_map: return ext_name_map[c_name]
    
    return {'math': False, 'sci': False, 'eng': False, 'gen': False}

def process_student_data(s_rec, default_grade):
    sid = s_rec['sid']
    rank_str = f"{s_rec['rank']}/{s_rec['total']} ({s_rec['group_label']})"
    start_g, start_s = determine_start_grade(s_rec)
    
    my_scores = df_scores[df_scores[sc_id_col] == sid].copy()
    if my_scores.empty:
        return {}, {'數學': 0, '基礎科學': 0, '工程專業': 0, '通識': 0, '總計': 0}, rank_str, sid
        
    my_scores = my_scores.sort_values(by=[sc_ay_col, sc_sem_col])
    valid_years = [to_int(y) for y in my_scores[sc_ay_col].unique() if to_int(y) > 0]
    base_year = min(valid_years) if valid_years else 0
    
    sem_data = {} 
    total_stats = {'數學': 0, '基礎科學': 0, '工程專業': 0, '通識': 0, '總計': 0}
    
    for _, row in my_scores.iterrows():
        raw_score = row[sc_score_col]
        try: score_val = float(raw_score)
        except: score_val = -1 
        
        is_valid = True
        if score_val != -1:
            if score_val == 999: is_valid = False
            if score_val < 60: is_valid = False
        if not is_valid: continue

        ay = to_int(row[sc_ay_col])
        sem = to_int(row[sc_sem_col])
        code = row[sc_code_col]
        name = row[sc_name_col]
        try: cred = float(row[sc_credit_col])
        except: cred = 0
            
        if base_year > 0 and ay > 0:
            diff = ay - base_year
            current_grade = start_g + diff
        else:
            current_grade = 0
            
        sem_key_tuple = (current_grade, sem)
        if sem_key_tuple not in sem_data:
            sem_data[sem_key_tuple] = {'courses': [], 'stats': {'數學': 0, '基礎科學': 0, '工程專業': 0, '通識': 0, '總計': 0}}
        
        cats = get_course_category_dual(name, code)
        
        if cats['math']: 
            sem_data[sem_key_tuple]['stats']['數學'] += cred
            total_stats['數學'] += cred
        if cats['sci']: 
            sem_data[sem_key_tuple]['stats']['基礎科學'] += cred
            total_stats['基礎科學'] += cred
        if cats['eng']: 
            sem_data[sem_key_tuple]['stats']['工程專業'] += cred
            total_stats['工程專業'] += cred
        if cats['gen']: 
            sem_data[sem_key_tuple]['stats']['通識'] += cred
            total_stats['通識'] += cred
            
        sem_data[sem_key_tuple]['stats']['總計'] += cred
        total_stats['總計'] += cred
        
        if not any(cats.values()):
            if code not in db_code_map and name not in db_name_map and code not in ext_code_map and name not in ext_name_map:
                missing_courses.add(f"{code} {name}")
        
        try:
            is_int = float(score_val).is_integer()
        except:
            is_int = False
        score_display = int(score_val) if score_val != -1 and is_int else raw_score
        
        try:
            cred_is_int = float(cred).is_integer()
        except:
            cred_is_int = False
        cred_display = int(cred) if cred_is_int else cred
        
        c_fmt = f"{name}({code}, {score_display}, {cred_display})"
        sem_data[sem_key_tuple]['courses'].append(c_fmt)
        
    return sem_data, total_stats, rank_str, sid


# ==========================================
# 7. 報表輸出與 IEET 自動化審查門檻模組
# ==========================================
def write_report(data, filename, default_grade):
    # 💡 [全新優化] 呼叫 get_unique_filepath 自動檢查檔案是否存在，防止同天重複執行時覆蓋成果
    full_path = get_unique_filepath(OUTPUT_DIR_PATH, filename)
    print(f"正在建置與導出 Excel 審查報表: {full_path}")
    
    is_undergrad = "大學部" in filename
    MIN_GRAD_CREDITS = MIN_GRAD_CREDITS_UNDERGRAD if is_undergrad else MIN_GRAD_CREDITS_GRAD 
    
    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
        for year in sorted(data.keys()):
            sheet_name = f"{year}學年"
            ws = writer.book.create_sheet(sheet_name)
            
            cursor = 1
            for s_rec in data[year]:
                sem_data, total_stats, rank_str, sid = process_student_data(s_rec, default_grade)
                
                header = f"學號: {sid}  |  排名: {rank_str}  |  最低畢業學分基準: {MIN_GRAD_CREDITS}"
                ws.cell(row=cursor, column=1, value=header).font = Font(bold=True, size=12)
                cursor += 1
                
                headers = ["年級(學期)", "課程(課號、成績、學分)", "數學", "基礎科學", "工程專業", "通識", "學期總計"]
                for i, h in enumerate(headers, 1):
                    cell = ws.cell(row=cursor, column=i, value=h)
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                cursor += 1
                
                def sort_key(k):
                    try: g, s = k; return (g, s)
                    except: return (99, 99)
                
                sorted_keys = sorted(sem_data.keys(), key=sort_key)
                has_capstone = False 
                
                for k in sorted_keys:
                    g, s = k
                    display_sem = format_semester_key(f"{g}({s})")
                    info = sem_data[k]
                    courses_str = "、".join(info['courses'])
                    stats = info['stats']
                    
                    for c_fmt in info['courses']:
                        if any(kw in c_fmt for kw in ['專題', '設計', '實作']):
                            has_capstone = True

                    ws.cell(row=cursor, column=1, value=display_sem).alignment = Alignment(horizontal='center')
                    ws.cell(row=cursor, column=2, value=courses_str).alignment = Alignment(wrap_text=True, vertical='center')
                    
                    cat_keys = ['數學', '基礎科學', '工程專業', '通識', '總計']
                    for i, key in enumerate(cat_keys, 3):
                        val = stats[key]
                        try:
                            val_is_int = float(val).is_integer()
                        except:
                            val_is_int = False
                        v_str = int(val) if val_is_int else val
                        cell = ws.cell(row=cursor, column=i, value=v_str if v_str > 0 else "")
                        cell.alignment = Alignment(horizontal='center')
                        
                    for i in range(1, 8):
                        ws.cell(row=cursor, column=i).border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                                                     top=Side(style='thin'), bottom=Side(style='thin'))
                    cursor += 1
                
                ws.cell(row=cursor, column=1, value="累積總學分").font = Font(bold=True)
                ws.cell(row=cursor, column=1).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                ws.cell(row=cursor, column=2).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                
                cat_keys = ['數學', '基礎科學', '工程專業', '通識', '總計']
                for i, key in enumerate(cat_keys, 3):
                    val = total_stats[key]
                    try:
                        val_is_int = float(val).is_integer()
                    except:
                        val_is_int = False
                    v_str = int(val) if val_is_int else val
                    cell = ws.cell(row=cursor, column=i, value=v_str)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                
                for i in range(1, 8):
                    ws.cell(row=cursor, column=i).border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                                                 top=Side(style='thin'), bottom=Side(style='thin'))
                cursor += 1

                if is_undergrad:
                    math_c = total_stats['數學']
                    sci_c = total_stats['基礎科學']
                    eng_c = total_stats['工程專業']
                    
                    is_411_ok = (math_c >= 9) and (sci_c >= 9) and ((math_c + sci_c) >= (MIN_GRAD_CREDITS * 0.25))
                    res_411 = "[符合]" if is_411_ok else "[不符合]"
                    detail_411 = f"{res_411} 4.1.1 數學及基礎科學 (各>=9且合計>=畢業基數{MIN_GRAD_CREDITS}之1/4)。 詳情: 數:{int(math_c)}, 理:{int(sci_c)}, 合計佔比:{calc_pct(math_c+sci_c, MIN_GRAD_CREDITS)}"
                    
                    ws.cell(row=cursor, column=1, value="IEET 4.1.1").font = Font(bold=True)
                    ws.merge_cells(start_row=cursor, start_column=2, end_row=cursor, end_column=7)
                    cell_411 = ws.cell(row=cursor, column=2, value=detail_411)
                    if not is_411_ok: cell_411.font = Font(color="FF0000", bold=True)
                    cursor += 1
                    
                    is_412_ok = (eng_c >= (MIN_GRAD_CREDITS * 0.375)) and has_capstone
                    res_412 = "[符合]" if is_412_ok else "[不符合]"
                    detail_412 = f"{res_412} 4.1.2 工程專業 (>=畢業基數{MIN_GRAD_CREDITS}之3/8 且含專題)。 詳情: 佔比:{calc_pct(eng_c, MIN_GRAD_CREDITS)}, 專題:{'有' if has_capstone else '無'}"
                    
                    ws.cell(row=cursor, column=1, value="IEET 4.1.2").font = Font(bold=True)
                    ws.merge_cells(start_row=cursor, start_column=2, end_row=cursor, end_column=7)
                    cell_412 = ws.cell(row=cursor, column=2, value=detail_412)
                    if not is_412_ok: cell_412.font = Font(color="FF0000", bold=True)
                    cursor += 1

                cursor += 2
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 85
            for col in ['C', 'D', 'E', 'F', 'G']:
                ws.column_dimensions[col].width = 11
            
        if 'Sheet' in writer.book.sheetnames: writer.book.remove(writer.book['Sheet'])


# ==========================================
# 8. 結束處理 (主入口執行程序)
# ==========================================
write_report(u_data, f"大學部_畢業生成績分析_{today_str}.xlsx", 4)
write_report(g_data, f"碩士班_畢業生成績分析_{today_str}.xlsx", 2)

if missing_courses:
    with open(os.path.join(OUTPUT_DIR_PATH, "missing_courses.txt"), "w", encoding='utf-8') as f:
        f.write("\n".join(sorted(list(missing_courses))))
    print(f"注意：有 {len(missing_courses)} 門課程於資料庫與備援表中皆未被分類，詳見 missing_courses.txt")

db.close()
print("完成！")