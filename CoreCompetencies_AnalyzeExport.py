import pandas as pd
from Accessdb import AccessHelper
import os
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ==========================================
# 1. 設定與準備
# ==========================================
db = AccessHelper()
today_str = datetime.today().strftime('%Y%m%d')

# [設定] 檔案輸出路徑
BASE_DIR = 'output_files'
SUB_DIR = '核心能力與教育目標分析'
OUTPUT_DIR_PATH = os.path.join(BASE_DIR, SUB_DIR)

# 自動建立資料夾
if not os.path.exists(OUTPUT_DIR_PATH):
    os.makedirs(OUTPUT_DIR_PATH)
    print(f"已建立資料夾: {OUTPUT_DIR_PATH}")

# 定義核心能力 (K1-K5)
core_competencies = {
    'K1': '能夠整合、組織電機專業理論來分析、表達問題之能力。',
    'K2': '能夠運用電機專業知識解決及實作電機工程問題之能力。',
    'K3': '具備分工、協調、重視團隊合作精神、遵守工程倫理以達成工作目標之能力。',
    'K4': '能夠激發自己潛能、融合他人智慧，具備獨立思考以及研究創新之能力。',
    'K5': '具備吸收電機新知、掌握國際發展趨勢，隨時接受競爭挑戰之能力。'
}

# 定義教育目標與核心能力的對應關係 (計算公式)
peo_definitions = {
    '學識理論': ['K1', 'K2', 'K4', 'K5'],
    '專業技術': ['K1', 'K2', 'K3'],
    '團隊精神與工程倫理': ['K3', 'K4'],
    '獨立思考與研究創新': ['K1', 'K2', 'K3', 'K4', 'K5'],
    '國際視野': ['K4', 'K5']
}

# 資料庫欄位對應
db_col_map = {
    'K1': 'has_SO_K1',
    'K2': 'has_SO_K2',
    'K3': 'has_SO_K3',
    'K4': 'has_SO_K4',
    'K5': 'has_SO_K5'
}

print("正在讀取並連結資料庫 (Course_Matrix + Courses)...")

# [關鍵修正] 使用 JOIN 取得 dept_code
sql = """
SELECT 
    M.*, 
    C.dept_code 
FROM Course_Matrix AS M
INNER JOIN Courses AS C ON M.course_id = C.id
WHERE M.course_score_AVG IS NOT NULL
"""
df_matrix_all = pd.read_sql(sql, db.conn)

# 清理 dept_code 資料 (去除空白)
if not df_matrix_all.empty:
    df_matrix_all['dept_code'] = df_matrix_all['dept_code'].astype(str).str.strip()

# ==========================================
# 2. Excel 寫入核心邏輯
# ==========================================
def write_semester_sheet(ws, df_sem, sem_label):
    """
    撰寫單一學期的統計資料 (包含 K 與 PEO)
    回傳: 該學期的統計數據 dict (供趨勢表使用)
    """
    stats_result = {}
    
    # --- Part 1: 核心能力 (K1-K5) ---
    headers = ['核心能力', '對應課程與平均分數', '核心能力評量結果 (平均)']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    current_row = 2
    k_scores = {} 

    for k_key, k_desc in core_competencies.items():
        db_col = db_col_map[k_key]
        
        # 篩選課程
        df_k = df_sem[df_sem[db_col].apply(lambda x: bool(x) if pd.notnull(x) else False)].copy()
        
        # 收集課程與分數
        total_score = 0
        count = 0
        course_items = []
        
        df_k = df_k.sort_values('course_code')
        
        for _, row in df_k.iterrows():
            c_name = str(row['course_name']).strip()
            c_code = str(row['course_code']).strip()
            
            if pd.notnull(row['course_score_AVG']):
                score = float(row['course_score_AVG'])
                course_items.append(f"{c_name} {c_code}[{score:.1f}]")
                total_score += score
                count += 1
        
        course_list_str = ", ".join(course_items)
        
        avg_result = 0.0
        if count > 0:
            avg_result = round(total_score / count, 2)
        
        k_scores[k_key] = avg_result
        stats_result[k_key] = avg_result
        
        # 寫入 Excel
        cell_desc = ws.cell(row=current_row, column=1, value=f"{k_key} {k_desc}")
        cell_desc.alignment = Alignment(wrap_text=True, vertical='top')
        
        cell_courses = ws.cell(row=current_row, column=2, value=course_list_str)
        cell_courses.alignment = Alignment(wrap_text=True, vertical='top')
        
        cell_score = ws.cell(row=current_row, column=3, value=avg_result)
        cell_score.alignment = Alignment(horizontal='center', vertical='top')
        
        current_row += 1

    # --- Part 2: 教育目標 (PEOs) ---
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    title_cell = ws.cell(row=current_row, column=1, value="【教育目標達成度分析】")
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.font = Font(bold=True)
    title_cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    
    current_row += 1
    
    peo_headers = ['教育目標名稱', '計算公式 (核心能力平均)', '目標達成評量結果']
    for col_idx, header in enumerate(peo_headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    current_row += 1
    
    for peo_name, required_ks in peo_definitions.items():
        current_k_values = [k_scores.get(k, 0.0) for k in required_ks]
        
        if len(current_k_values) > 0:
            peo_avg = sum(current_k_values) / len(current_k_values)
            peo_avg = round(peo_avg, 2)
        else:
            peo_avg = 0.0
        
        stats_result[peo_name] = peo_avg
            
        formula_str = f"({' + '.join(required_ks)}) / {len(required_ks)}"
        
        ws.cell(row=current_row, column=1, value=peo_name)
        cell_formula = ws.cell(row=current_row, column=2, value=formula_str)
        cell_formula.alignment = Alignment(horizontal='center')
        cell_score = ws.cell(row=current_row, column=3, value=peo_avg)
        cell_score.alignment = Alignment(horizontal='center')
        
        current_row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 25
    
    return stats_result

def write_trend_sheet(ws, trend_data):
    """
    建立歷學期趨勢統計表 (拆分為兩個表格：核心能力 與 教育目標)
    """
    if not trend_data:
        return

    # --- 表格 1: 核心能力 ---
    current_row = 1
    
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    title_k = ws.cell(row=current_row, column=1, value="【歷學期核心能力評量結果趨勢】")
    title_k.font = Font(bold=True, size=14)
    title_k.alignment = Alignment(horizontal='center')
    title_k.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    current_row += 1
    headers_k = ['學年學期'] + [f"{k} {desc}" for k, desc in core_competencies.items()]
    
    for col_idx, header in enumerate(headers_k, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        if col_idx == 1:
            ws.column_dimensions[chr(64+col_idx)].width = 15
        else:
            ws.column_dimensions[chr(64+col_idx)].width = 30
            
    current_row += 1
    
    for row_data in trend_data:
        cell = ws.cell(row=current_row, column=1, value=row_data.get('Semester'))
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for i, key in enumerate(core_competencies.keys(), 2):
            val = row_data.get(key, 0.0)
            cell = ws.cell(row=current_row, column=i, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = '0.00'
        current_row += 1

    # --- 表格 2: 教育目標 ---
    current_row += 2
    
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    title_peo = ws.cell(row=current_row, column=1, value="【歷學期教育目標達成度趨勢】")
    title_peo.font = Font(bold=True, size=14)
    title_peo.alignment = Alignment(horizontal='center')
    title_peo.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    current_row += 1
    headers_peo = ['學年學期'] + list(peo_definitions.keys())
    
    for col_idx, header in enumerate(headers_peo, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
    current_row += 1
    
    for row_data in trend_data:
        cell = ws.cell(row=current_row, column=1, value=row_data.get('Semester'))
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for i, key in enumerate(peo_definitions.keys(), 2):
            val = row_data.get(key, 0.0)
            cell = ws.cell(row=current_row, column=i, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = '0.00'
        current_row += 1

# ==========================================
# 3. 匯出封裝函式
# ==========================================
def export_data(df, filename_prefix):
    output_filename = f'{filename_prefix}_核心能力與教育目標達成度分析_{today_str}.xlsx'
    full_output_path = os.path.join(OUTPUT_DIR_PATH, output_filename)
    
    if df.empty:
        print(f"警告：{filename_prefix} 沒有資料，略過匯出。")
        return

    print(f"準備寫入檔案: {full_output_path}")

    with pd.ExcelWriter(full_output_path, engine='openpyxl') as writer:
        
        # 排序與迭代
        df['sort_key'] = df['academic_year'] * 10 + df['semester']
        unique_sems = df[['academic_year', 'semester', 'sort_key']].drop_duplicates().sort_values('sort_key')
        
        trend_records = [] 
        
        for _, row in unique_sems.iterrows():
            year = int(row['academic_year'])
            sem = int(row['semester'])
            
            # 排除第3學期
            if sem == 3:
                continue
            
            sheet_name = f"{year}-{sem}"
            
            # 篩選學期資料
            df_current_sem = df[
                (df['academic_year'] == year) & 
                (df['semester'] == sem)
            ]
            
            # 建立學期分頁
            ws = writer.book.create_sheet(sheet_name)
            stats = write_semester_sheet(ws, df_current_sem, sheet_name)
            
            stats['Semester'] = sheet_name
            trend_records.append(stats)
            
        # 建立趨勢分頁
        if trend_records:
            ws_trend = writer.book.create_sheet("歷學期趨勢分析", 0)
            write_trend_sheet(ws_trend, trend_records)
            
        if 'Sheet' in writer.book.sheetnames:
            writer.book.remove(writer.book['Sheet'])

# ==========================================
# 4. 主程式執行
# ==========================================
if not df_matrix_all.empty:
    # 大學部 (B301)
    print("--- 正在處理大學部資料 ---")
    df_undergrad = df_matrix_all[df_matrix_all['dept_code'] == 'B301'].copy()
    export_data(df_undergrad, "大學部")
    
    # 碩士班 (M301)
    print("--- 正在處理碩士班資料 ---")
    df_grad = df_matrix_all[df_matrix_all['dept_code'] == 'M301'].copy()
    export_data(df_grad, "碩士班")

else:
    print("錯誤：資料庫中沒有有效的課程矩陣資料。")

db.close()
print("-" * 30)
print(f"全部完成！請檢查資料夾: {OUTPUT_DIR_PATH}")