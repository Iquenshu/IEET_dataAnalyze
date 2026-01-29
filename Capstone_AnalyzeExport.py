import pandas as pd
from Accessdb import AccessHelper
import os
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ==========================================
# 1. 設定與準備
# ==========================================
db = AccessHelper()
today_str = datetime.today().strftime('%Y%m%d')

# [設定] 檔案輸出路徑
BASE_DIR = 'output_files'
SUB_DIR = '必修專題課核心能力分析'
OUTPUT_DIR_PATH = os.path.join(BASE_DIR, SUB_DIR)

# 自動建立資料夾
if not os.path.exists(OUTPUT_DIR_PATH):
    os.makedirs(OUTPUT_DIR_PATH)
    print(f"已建立資料夾: {OUTPUT_DIR_PATH}")

# 定義核心能力 (K1-K5)
core_competencies = {
    'K1': '能夠整合、組織電機專業理論來分析、表達問題之能力。',
    'K2': '能夠運用電機專業知識解決及實作電機工程問題之能力。',
    'K3': '具備分工、協調、重視團隊合作精神、遵守工程倫理以達成工作目標之能力。',
    'K4': '能夠激發自己潛能、融合他人智慧，具備獨立思考以及研究創新之能力。',
    'K5': '具備吸收電機新知、掌握國際發展趨勢，隨時接受競爭挑戰之能力。'
}

# 資料庫欄位對應
db_col_map = {
    'K1': 'has_SO_K1',
    'K2': 'has_SO_K2',
    'K3': 'has_SO_K3',
    'K4': 'has_SO_K4',
    'K5': 'has_SO_K5'
}

print("正在讀取並連結資料庫 (Course_Matrix + Courses)...")

# 1. 取得所有課程詳細資料 (連結 Courses 以取得 dept_code, is_required 等)
sql = '''
SELECT 
    M.*, 
    C.dept_code, 
    C.is_required, 
    C.course_name as c_name_check 
FROM Course_Matrix AS M
INNER JOIN Courses AS C ON M.course_id = C.id
WHERE M.course_score_AVG IS NOT NULL
'''
df_matrix_all = pd.read_sql(sql, db.conn)

# ==========================================
# 2. 資料篩選與計算邏輯
# ==========================================

def filter_capstone_courses(df):
    """
    篩選「專題」且「必修」的課程
    1. course_name 包含 '專題'
    2. is_required 為 True (1)
    """
    # 確保字串欄位
    df['course_name'] = df['course_name'].astype(str)
    
    # 條件 1: 名稱含 "專題"
    mask_name = df['course_name'].str.contains('專題', na=False)
    
    # 條件 2: 必修 (Access 可能回傳 1/True)
    # 轉布林值比較保險
    mask_req = df['is_required'].apply(lambda x: bool(x) if pd.notnull(x) else False)
    
    return df[mask_name & mask_req].copy()

def calculate_k_avg(df_filtered):
    """
    計算 K1~K5 的平均分數 (針對篩選後的課程)
    回傳 dict: {'K1': 80.5, ... 'Total_Avg': 82.0, 'Valid_Count': 3}
    """
    results = {}
    valid_scores = []
    
    for k_key in ['K1', 'K2', 'K3', 'K4', 'K5']:
        db_col = db_col_map[k_key]
        
        # 找出該 K 被勾選的課程
        mask_k = df_filtered[db_col].apply(lambda x: bool(x) if pd.notnull(x) else False)
        df_k = df_filtered[mask_k]
        
        if not df_k.empty:
            avg_score = df_k['course_score_AVG'].mean()
            results[k_key] = round(avg_score, 2)
            valid_scores.append(avg_score)
        else:
            results[k_key] = 0.0
            
    # [修正邏輯] 計算有效項目的平均
    if valid_scores:
        total_avg = sum(valid_scores) / len(valid_scores)
        results['Total_Avg'] = round(total_avg, 2)
        results['Valid_Count'] = len(valid_scores) # 記錄有效項數
    else:
        results['Total_Avg'] = 0.0
        results['Valid_Count'] = 0
        
    return results

def get_course_list_str(df_filtered, k_key):
    """
    取得該 K 能力下的課程清單字串 (用於 Excel 顯示)
    """
    db_col = db_col_map[k_key]
    mask_k = df_filtered[db_col].apply(lambda x: bool(x) if pd.notnull(x) else False)
    df_k = df_filtered[mask_k].sort_values('course_code')
    
    items = []
    for _, row in df_k.iterrows():
        c_name = str(row['course_name']).strip()
        c_code = str(row['course_code']).strip()
        score = float(row['course_score_AVG'])
        items.append(f"{c_name} {c_code}[{score:.1f}]")
        
    return ", ".join(items)

# ==========================================
# 3. Excel 輸出邏輯
# ==========================================

def write_analysis_sheet(ws, df_data, title):
    """
    撰寫核心能力分析表 (單一學年或總合)
    """
    # 標題
    ws.merge_cells('A1:C1')
    cell_title = ws.cell(row=1, column=1, value=title)
    cell_title.font = Font(bold=True, size=14)
    cell_title.alignment = Alignment(horizontal='center')
    cell_title.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # 表頭
    headers = ['核心能力', '對應課程與平均分數', '評量結果 (平均)']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
    # 設定欄寬
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 20
    
    # 計算分數
    k_scores = calculate_k_avg(df_data)
    
    current_row = 3
    # 寫入 K1~K5
    for k_key, k_desc in core_competencies.items():
        # 第一欄: 核心能力描述
        ws.cell(row=current_row, column=1, value=f"{k_key} {k_desc}").alignment = Alignment(wrap_text=True, vertical='top')
        
        # 第二欄: 課程清單
        course_str = get_course_list_str(df_data, k_key)
        ws.cell(row=current_row, column=2, value=course_str).alignment = Alignment(wrap_text=True, vertical='top')
        
        # 第三欄: 平均分數
        val = k_scores[k_key]
        cell = ws.cell(row=current_row, column=3, value=val)
        cell.alignment = Alignment(horizontal='center', vertical='top')
        
        current_row += 1
        
    # [修正] 寫入總平均列與說明
    label_text = f"核心能力總平均 (排除未評量項目)"
    ws.cell(row=current_row, column=1, value=label_text).font = Font(bold=True)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
    
    # 合併第 1, 2 欄
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    
    # 總平均分數
    total_val = k_scores['Total_Avg']
    cell_total = ws.cell(row=current_row, column=3, value=total_val)
    cell_total.font = Font(bold=True, color="FF0000") # 紅字粗體
    cell_total.alignment = Alignment(horizontal='center')
    
    # [新增] 底部備註
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    note_text = "* 備註：總平均計算僅包含有對應到課程之核心能力項目 (分數 > 0)，未對應項目不列入分母計算。"
    cell_note = ws.cell(row=current_row, column=1, value=note_text)
    cell_note.font = Font(italic=True, size=10, color="555555")
    cell_note.alignment = Alignment(horizontal='left')
    
    return k_scores 

def write_trend_sheet(ws, trend_data):
    """
    歷年趨勢總表 (學年 vs K1~K5 + 有效平均)
    """
    ws.merge_cells('A1:G1')
    cell_title = ws.cell(row=1, column=1, value="【歷年必修專題課程核心能力趨勢分析】")
    cell_title.font = Font(bold=True, size=14)
    cell_title.alignment = Alignment(horizontal='center')
    cell_title.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # [修正] 欄位名稱
    headers = ['學年', 'K1', 'K2', 'K3', 'K4', 'K5', '有效平均']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    current_row = 3
    for row_data in trend_data:
        ws.cell(row=current_row, column=1, value=row_data['Year']).alignment = Alignment(horizontal='center')
        
        for i, k in enumerate(['K1', 'K2', 'K3', 'K4', 'K5', 'Total_Avg'], 2):
            val = row_data.get(k, 0.0)
            ws.cell(row=current_row, column=i, value=val).alignment = Alignment(horizontal='center')
            
        current_row += 1
        
    # [新增] 底部備註
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    note_text = "* 備註：「有效平均」僅計算該學年度有開課並評分之核心能力項目。"
    cell_note = ws.cell(row=current_row, column=1, value=note_text)
    cell_note.font = Font(italic=True, size=10, color="555555")

# ==========================================
# 4. 主流程
# ==========================================
def process_and_export(df_all, dept_name):
    print(f"\n--- 正在處理 {dept_name} 資料 ---")
    
    # 1. 篩選專題且必修
    df_capstone = filter_capstone_courses(df_all)
    
    if df_capstone.empty:
        print(f"警告：找不到 {dept_name} 的必修專題課程。")
        return

    print(f"共找到 {len(df_capstone)} 門必修專題課程。")
    
    # 輸出檔案設定
    output_filename = f'{dept_name}_必修專題核心能力分析_{today_str}.xlsx'
    full_path = os.path.join(OUTPUT_DIR_PATH, output_filename)
    
    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
        
        trend_records = []
        
        # --- 分頁 1: 全部學年總合 ---
        ws_all = writer.book.create_sheet("全部學年總合")
        write_analysis_sheet(ws_all, df_capstone, f"【{dept_name} 歷年必修專題課程總合分析】")
        
        # --- 分頁 2~N: 各學年分析 ---
        years = sorted(df_capstone['academic_year'].unique())
        
        for year in years:
            sheet_name = f"{year}學年度"
            df_year = df_capstone[df_capstone['academic_year'] == year]
            
            ws_year = writer.book.create_sheet(sheet_name)
            scores = write_analysis_sheet(ws_year, df_year, f"【{dept_name} {year}學年度 必修專題課程分析】")
            
            # 收集趨勢資料
            scores['Year'] = f"{year}學年度"
            trend_records.append(scores)
            
        # --- 分頁 0: 歷年趨勢表 (插在最前面) ---
        if trend_records:
            ws_trend = writer.book.create_sheet("歷年趨勢統計", 0)
            write_trend_sheet(ws_trend, trend_records)
            
        if 'Sheet' in writer.book.sheetnames:
            writer.book.remove(writer.book['Sheet'])
            
    print(f"輸出完成: {full_path}")

# ==========================================
# 5. 執行
# ==========================================
if not df_matrix_all.empty:
    # 資料清理
    df_matrix_all['dept_code'] = df_matrix_all['dept_code'].astype(str).str.strip()
    
    # 分別處理大學部與碩士班
    # 大學部 (B301)
    df_undergrad = df_matrix_all[df_matrix_all['dept_code'] == 'B301'].copy()
    process_and_export(df_undergrad, "大學部")
    
    # 碩士班 (M301)
    df_grad = df_matrix_all[df_matrix_all['dept_code'] == 'M301'].copy()
    process_and_export(df_grad, "碩士班")
else:
    print("錯誤：資料庫中沒有課程矩陣資料。")

db.close()