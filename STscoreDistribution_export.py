import pandas as pd
from Accessdb import AccessHelper
import os
from datetime import datetime
from openpyxl.styles import Font
from openpyxl import load_workbook

# 1. 連接 Access 資料庫
db = AccessHelper()

# 2. 讀取 STscoreAnalyze 資料表（只取分數區間，不含 total）
sql = "SELECT 學年度, 課號, 課程名稱, 分數區間, 人數 FROM STscoreAnalyze WHERE 分數區間 <> 'total'"
df = pd.read_sql(sql, db.conn)
db.close()

# 統一課號與課程名稱格式
df['課號'] = df['課號'].astype(str).str.strip()
df['課程名稱'] = df['課程名稱'].astype(str).str.strip()

# 3. 準備輸出路徑
output_dir = 'output_files'
os.makedirs(output_dir, exist_ok=True)
today_str = datetime.today().strftime('%Y%m%d')
output_path = os.path.join(output_dir, f'STscoreDistributionData_{today_str}.xlsx')

score_labels = ["0-9","10-19","20-29","30-39","40-49","50-59","60-69","70-79","80-89","90-100"]
all_years = sorted(df['學年度'].unique())

# 4. 整理所有課程的分數分段分布，合併到同一分頁
output_rows = []
for (course_name, course_id), group in df.groupby(['課程名稱', '課號']):
    # 該課程有資料的學年度
    years = sorted(group['學年度'].unique())
    # 第一列：課程名稱
    output_rows.append([f"課程名稱: {course_name}"])
    # 第二列：課號
    output_rows.append([f"課號: {course_id}"])
    # 第三列：學年度橫向標題
    output_rows.append(['分數區間'] + years)
    # 依分數區間排序
    for label in score_labels:
        row = [label]
        for year in years:
            count = group[(group['學年度'] == year) & (group['分數區間'] == label)]['人數'].sum()
            row.append(int(count))
        output_rows.append(row)
    # 最下面一列加上總修課人數
    total_row = ['總修課人數']
    for year in years:
        total = group[group['學年度'] == year]['人數'].sum()
        total_row.append(int(total))
    output_rows.append(total_row)
    # 空行分隔不同課程
    output_rows.append([])

# 5. 輸出到 Excel
out_df = pd.DataFrame(output_rows)
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    out_df.to_excel(writer, sheet_name='所有課程分布', index=False, header=False)

# 6. 加粗課程名稱與課號
wb = load_workbook(output_path)
ws = wb['所有課程分布']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row)):
    # 第一列：課程名稱
    if row[0].value and str(row[0].value).startswith("課程名稱:"):
        for cell in row:
            cell.font = Font(bold=True)
        # 下一列：課號也加粗
        if i + 1 < ws.max_row:
            for cell in ws[i + 2]:
                cell.font = Font(bold=True)
wb.save(output_path)

print(f"資料已匯出至 {output_path}")