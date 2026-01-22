import pandas as pd
from Accessdb import AccessHelper
import os
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows

# ==========================================
# 1. 設定與準備
# ==========================================
db = AccessHelper()
today_str = datetime.today().strftime('%Y%m%d')

# [設定] 檔案輸出路徑
BASE_DIR = 'output_files'
SUB_DIR = '離系問券統計'
OUTPUT_DIR_PATH = os.path.join(BASE_DIR, SUB_DIR)

# 自動建立資料夾
if not os.path.exists(OUTPUT_DIR_PATH):
    os.makedirs(OUTPUT_DIR_PATH)
    print(f"已建立資料夾: {OUTPUT_DIR_PATH}")

print("正在讀取資料庫...")
df_u = pd.read_sql("SELECT * FROM LDUdataAnalyze", db.conn)
df_g = pd.read_sql("SELECT * FROM LDGdataAnalyze", db.conn)
df_q = pd.read_sql("SELECT * FROM LeavDepQuest", db.conn)

def get_quest_map(row):
    """解析問卷題目對照表"""
    big_map = {}
    small_map = {}
    for i in range(1, 4):
        big_key = f"A{i}0"
        if big_key in row:
            big_map[big_key] = row[big_key]
        for j in range(1, 12):
            small_key = f"A{i}{j}"
            if small_key in row:
                small_map[small_key] = row[small_key]
    return big_map, small_map

# 取得題目對照
quest_row_u = df_q[df_q['QuestType'] == 'UdataZH'].iloc[0]
quest_row_g = df_q[df_q['QuestType'] == 'GdataZH'].iloc[0]
big_map_u, small_map_u = get_quest_map(quest_row_u)
big_map_g, small_map_g = get_quest_map(quest_row_g)

# ==========================================
# 2. Excel 寫入核心邏輯
# ==========================================

def write_sheet_data(ws, df_source, big_map, small_map, startrow=0, label="", show_pct=False):
    """
    寫入標準統計表
    """
    ws.cell(row=startrow+1, column=1, value=label)
    startrow += 1
    
    # 定義表頭
    if show_pct:
        header = [
            '題目', 
            '完全未達成', '完全未達成佔比', 
            '小部分達成', '小部分達成佔比', 
            '部分達成', '部分達成佔比', 
            '大部分達成', '大部分達成佔比', 
            '完全達成', '完全達成佔比', 
            '總數'
        ]
    else:
        header = ['題目', '完全未達成', '小部分達成', '部分達成', '大部分達成', '完全達成', '總數']

    # 遍歷大題
    for big_key, big_title in big_map.items():
        if big_key == "A30": continue
            
        ws.cell(row=startrow+1, column=1, value=f"【{big_title}】")
        startrow += 1
        
        qids = [k for k in small_map if k.startswith(big_key[:2]) and not k.startswith("A30")]
        
        table_data = []
        # [修改 1] 使用 enumerate 加入題號 (1. xxx)
        for idx, qid in enumerate(qids, 1):
            row = df_source[df_source['qid'] == qid]
            if row.empty: continue
            
            row_data = row.iloc[0]
            total = row_data['total'] if row_data['total'] > 0 else 1 
            
            # 處理題目名稱：加入編號
            q_name = small_map.get(qid, qid)
            numbered_q_name = f"{idx}.{q_name}"
            
            col_data = [numbered_q_name]
            
            # 依序處理 5 個答案
            for i in range(1, 6):
                count_val = row_data[f'count_{i}']
                col_data.append(count_val)
                
                if show_pct:
                    if total > 0:
                        pct_val = round(count_val / total, 2)
                    else:
                        pct_val = 0.0
                    col_data.append(pct_val)
            
            col_data.append(row_data['total'])
            table_data.append(col_data)
            
        if table_data:
            table_df = pd.DataFrame(table_data, columns=header)
            
            # 寫入 Excel
            for r_idx, row in enumerate(dataframe_to_rows(table_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=startrow + r_idx, column=c_idx, value=value)
            
            startrow += len(table_data) + 3
            
    return startrow

def write_yearly_trend(ws, df_full, big_map, small_map, startrow):
    """
    寫入歷年變化趨勢分析表
    """
    ws.cell(row=startrow, column=1, value="【歷年變化趨勢分析 (109-113年)】")
    startrow += 2

    df_raw = df_full[~df_full['sem'].str.endswith('T')].copy()
    df_raw['year'] = df_raw['sem'].str[:3]
    years = sorted(list(set(df_raw['year'])))
    
    header = ['學年', '完全未達成佔比', '小部分達成佔比', '部分達成佔比', '大部分達成佔比', '完全達成佔比', '有效樣本數']

    for big_key, big_title in big_map.items():
        if big_key == "A30": continue
        
        ws.cell(row=startrow, column=1, value=f"類別：{big_title}")
        startrow += 1
        
        qids = [k for k in small_map if k.startswith(big_key[:2]) and not k.startswith("A30")]
        
        # [修改 1] 使用 enumerate 加入題號
        for idx, qid in enumerate(qids, 1):
            q_name = small_map.get(qid, qid)
            numbered_q_name = f"{idx}.{q_name}"
            
            ws.cell(row=startrow, column=1, value=f"題目：{numbered_q_name}")
            startrow += 1
            
            # 寫入表頭
            for c_idx, h_val in enumerate(header, 1):
                ws.cell(row=startrow, column=c_idx, value=h_val)
            startrow += 1
            
            # 寫入各學年數據
            for year in years:
                df_y = df_raw[(df_raw['qid'] == qid) & (df_raw['year'] == year)]
                
                if df_y.empty:
                    counts = {f'count_{i}': 0 for i in range(1, 6)}
                    counts['total'] = 0
                else:
                    counts = df_y[['count_1', 'count_2', 'count_3', 'count_4', 'count_5', 'total']].sum()
                
                total = counts['total']
                if total == 0: total = 1
                
                ws.cell(row=startrow, column=1, value=f"{year}學年")
                
                for i in range(1, 6):
                    val = round(counts[f'count_{i}'] / total, 2)
                    ws.cell(row=startrow, column=1+i, value=val)
                
                ws.cell(row=startrow, column=7, value=counts['total'])
                
                startrow += 1
            
            startrow += 1 
        
        startrow += 1 
        
    return startrow

def export_excel(df, big_map, small_map, filename, sheet_suffix):
    full_output_path = os.path.join(OUTPUT_DIR_PATH, filename)
    print(f"準備寫入檔案: {full_output_path}")

    with pd.ExcelWriter(full_output_path, engine='openpyxl') as writer:
        
        # --- [分頁 1] 全部學年統計 ---
        df_raw = df[~df['sem'].str.endswith('T')].copy()
        
        if not df_raw.empty:
            ws_all = writer.book.create_sheet("全部學年統計")
            
            # 總計表
            df_all_years_sum = df_raw.groupby('qid')[
                ['count_1', 'count_2', 'count_3', 'count_4', 'count_5', 'total']
            ].sum().reset_index()
            
            current_row = write_sheet_data(
                ws_all, 
                df_all_years_sum, 
                big_map, 
                small_map, 
                startrow=0, 
                label=f"歷年總計 (所有學年)",
                show_pct=True # 全部學年：開啟百分比
            )
            
            # 趨勢表
            current_row += 2
            write_yearly_trend(
                ws_all, 
                df, 
                big_map, 
                small_map, 
                startrow=current_row
            )

        # --- [分頁 2...] 各學年分頁 ---
        years = sorted(list(set(df['sem'].str[:3])))
        for year in years:
            sheet_name = f"{year}_{sheet_suffix}"
            ws = writer.book.create_sheet(sheet_name)
            current_row = 0
            
            # 學年總計
            df_year_raw = df[(df['sem'].str.startswith(year)) & (~df['sem'].str.endswith('T'))]
            if not df_year_raw.empty:
                df_year_total = df_year_raw.groupby('qid')[
                    ['count_1', 'count_2', 'count_3', 'count_4', 'count_5', 'total']
                ].sum().reset_index()
                
                # [修改 2] 這裡將 show_pct 改為 True，滿足您的需求
                current_row = write_sheet_data(
                    ws, 
                    df_year_total, 
                    big_map, 
                    small_map, 
                    startrow=current_row, 
                    label=f"{year}學年總和", 
                    show_pct=True # <--- 改為 True
                )
                current_row += 1
            
            # 學期明細 (維持不顯示百分比，避免版面過雜，若需要也可改 True)
            for sem_suffix in ["01", "02"]:
                sem = f"{year}{sem_suffix}"
                df_sem = df[df['sem'] == sem]
                if not df_sem.empty:
                    current_row = write_sheet_data(ws, df_sem, big_map, small_map, startrow=current_row, label=f"{sem}學期", show_pct=False)
                    current_row += 1

# ==========================================
# 3. 執行匯出
# ==========================================
print(f"輸出目標資料夾: {OUTPUT_DIR_PATH}")

print("正在匯出大學部資料...")
export_excel(df_u, big_map_u, small_map_u, f'大學部離系問券統計_{today_str}.xlsx', '大學部')

print("正在匯出研究所資料...")
export_excel(df_g, big_map_g, small_map_g, f'研究所離系問券統計_{today_str}.xlsx', '研究所')

db.close()
print("-" * 30)
print(f"執行完成！請檢查: {OUTPUT_DIR_PATH}")