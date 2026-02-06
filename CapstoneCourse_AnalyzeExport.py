import pandas as pd
from Accessdb import AccessHelper
import os
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import warnings

# 忽略警告
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)

# ==========================================
# 1. 設定與準備
# ==========================================
db = AccessHelper()
today_str = datetime.today().strftime('%Y%m%d')

BASE_DIR = 'output_files'
SUB_DIR = '必修專題課程個別分析'
OUTPUT_DIR_PATH = os.path.join(BASE_DIR, SUB_DIR)

if not os.path.exists(OUTPUT_DIR_PATH):
    os.makedirs(OUTPUT_DIR_PATH)
    print(f"已建立資料夾: {OUTPUT_DIR_PATH}")

# 核心能力完整標題
k_headers = {
    'K1': 'K1 能夠整合、組織電機專業理論來分析、表達問題之能力',
    'K2': 'K2 能夠運用電機專業知識解決及實作電機工程問題之能力',
    'K3': 'K3 具備分工、協調、重視團隊合作精神、遵守工程倫理以達成工作目標之能力',
    'K4': 'K4 能夠激發自己潛能、融合他人智慧，具備獨立思考以及研究創新之能力',
    'K5': 'K5 具備吸收電機新知、掌握國際發展趨勢，隨時接受競爭挑戰之能力'
}

# 評量方式對照表 (原始名稱)
smc_raw_mapping = {
    'smc_1': '紙筆測驗',
    'smc_2': '課堂討論',
    'smc_3': '個人書面報告或作品',
    'smc_4': '群組書面報告或作品',
    'smc_5': '個人口頭報告',
    'smc_6': '群組口頭報告',
    'smc_7': '校外參訪及實習',
    'smc_8': '證照/檢定',
    'smc_9': '活動及競賽',
    'smc_10': '課外閱讀'
}

# ==========================================
# 2. 資料讀取與前處理
# ==========================================

def get_capstone_courses():
    print("正在讀取課程資料 (Courses + Course_Matrix)...")
    sql = """
    SELECT 
        C.id as course_id,
        C.academic_year, C.semester, C.dept_code,
        C.course_code, C.course_name, C.instructor,
        C.is_required, C.credits,
        C.is_math, C.is_science, C.is_eng_prof,
        M.has_SO_K1, M.has_SO_K2, M.has_SO_K3, M.has_SO_K4, M.has_SO_K5,
        M.course_score_AVG as matrix_avg_score
    FROM Courses AS C
    LEFT JOIN Course_Matrix AS M ON C.id = M.course_id
    """
    df = pd.read_sql(sql, db.conn)
    
    # 必修判斷
    df['is_required_bool'] = df['is_required'].apply(lambda x: bool(x) if pd.notnull(x) else False)
    
    # 課名篩選
    mask_name = df['course_name'].astype(str).str.contains('專題', na=False)
    
    # 綜合篩選 (必修 + 專題)
    df_capstone = df[mask_name & df['is_required_bool']].copy()
    
    print(f"共篩選出 {len(df_capstone)} 門必修專題課程。")
    return df_capstone

def get_assessment_methods_formatted():
    """
    格式化評量方式 (■/□ 顯示)
    """
    print("正在讀取並整理評量方式...")
    sql = "SELECT * FROM Course_Competencies"
    df = pd.read_sql(sql, db.conn)
    
    if df.empty: return {}

    # 轉數值
    target_cols = list(smc_raw_mapping.keys())
    for col in target_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

    # 依 course_id 取最大值
    df_grouped = df.groupby('course_id')[target_cols].max().reset_index()
    
    assessment_dict = {}
    
    for _, row in df_grouped.iterrows():
        display_lines = []
        
        # 1. 紙筆測驗
        has_paper = (row['smc_1'] == 1)
        sym = "■" if has_paper else "□"
        display_lines.append(f"{sym} 紙筆測驗")
            
        # 2. 書面報告或作品 (合併 3, 4)
        has_written = (row['smc_3'] == 1 or row['smc_4'] == 1)
        sym = "■" if has_written else "□"
        display_lines.append(f"{sym} 書面報告或作品")
            
        # 3. 口頭報告 (合併 5, 6)
        has_oral = (row['smc_5'] == 1 or row['smc_6'] == 1)
        sym = "■" if has_oral else "□"
        display_lines.append(f"{sym} 口頭報告")
            
        # 4. 其他 (合併 2, 7, 8, 9, 10)
        others_details = []
        if row['smc_2'] == 1: others_details.append("課堂討論")
        if row['smc_7'] == 1: others_details.append("校外參訪及實習")
        if row['smc_8'] == 1: others_details.append("證照/檢定")
        if row['smc_9'] == 1: others_details.append("活動及競賽")
        if row['smc_10'] == 1: others_details.append("課外閱讀")
        
        if others_details:
            display_lines.append(f"■ 其他 ({'、'.join(others_details)})")
        else:
            display_lines.append("□ 其他")
            
        # 使用換行符號連接，讓 Excel 格子內分行顯示
        assessment_dict[row['course_id']] = "\n".join(display_lines)
        
    return assessment_dict

def get_student_stats():
    print("正在計算學生成績統計 (STscore)...")
    sql = "SELECT 學年度, 學期, 課號, 成績 FROM STscore"
    df = pd.read_sql(sql, db.conn)
    
    df['成績'] = pd.to_numeric(df['成績'], errors='coerce')
    df = df.dropna(subset=['成績'])
    
    # 排除退選 (999) 及異常值
    df_valid = df[(df['成績'] >= 0) & (df['成績'] <= 100)].copy()
    
    # Key 轉字串
    df_valid['學年度'] = df_valid['學年度'].astype(str)
    df_valid['學期'] = df_valid['學期'].astype(str)
    df_valid['課號'] = df_valid['課號'].astype(str).str.strip()
    
    stats = df_valid.groupby(['學年度', '學期', '課號']).agg(
        total_students=('成績', 'count'),
        pass_count=('成績', lambda x: (x >= 60).sum())
    ).reset_index()
    
    stats['pass_rate'] = (stats['pass_count'] / stats['total_students']) * 100
    stats['pass_rate'] = stats['pass_rate'].round(1)
    
    return stats

# ==========================================
# 3. 資料整合
# ==========================================

def determine_course_type(row):
    types = []
    if bool(row['is_math']): types.append("數學課程")
    if bool(row['is_science']): types.append("基礎科學課程")
    if bool(row['is_eng_prof']): types.append("工程專業課程")
    if not types: return "一般課程"
    return "、".join(types)

def check_k_association(row):
    for k in ['K1', 'K2', 'K3', 'K4', 'K5']:
        if bool(row[f'has_SO_{k}']):
            return True
    return False

def merge_data():
    df_main = get_capstone_courses()
    if df_main.empty: return pd.DataFrame()

    assess_dict = get_assessment_methods_formatted()
    df_stats = get_student_stats()
    
    final_rows = []
    
    for _, row in df_main.iterrows():
        c_id = row['course_id']
        ay = str(row['academic_year'])
        sem = str(row['semester'])
        code = str(row['course_code']).strip()
        dept = str(row['dept_code']).strip()
        
        # 1. 取得評量方式 (若無資料，預設全部 □)
        default_assess = "□ 紙筆測驗\n□ 書面報告或作品\n□ 口頭報告\n□ 其他"
        assess_str = assess_dict.get(c_id, default_assess)
        
        # 2. 取得學生成績統計
        stat_match = df_stats[
            (df_stats['學年度'] == ay) &
            (df_stats['學期'] == sem) &
            (df_stats['課號'] == code)
        ]
        
        if not stat_match.empty:
            s_row = stat_match.iloc[0]
            num_students = s_row['total_students']
            pass_rate = s_row['pass_rate']
        else:
            num_students = 0
            pass_rate = 0.0
            
        # 3. 取得平均成績 (Matrix)
        avg_score = row['matrix_avg_score'] if pd.notnull(row['matrix_avg_score']) else 0.0
        
        # 4. [篩選]
        has_core_comp = check_k_association(row)
        if num_students == 0 or not has_core_comp:
            continue
            
        # 建構資料列
        data = {
            '課號': code,
            '課程名稱': row['course_name'],
            '開課老師': row['instructor'],
            '學年': int(ay), # [新增]
            '開課學期': int(sem),
            '學分數': row['credits'],
            '必修/選修': "必修" if row['is_required_bool'] else "選修",
            '課程類型': determine_course_type(row),
            '修課人數': num_students,
            '評量方式': assess_str,
            '平均成績': round(avg_score, 1),
            '及格率(%)': pass_rate,
            'dept_code': dept
        }
        
        # 加入 K1~K5
        for k_key, k_title in k_headers.items():
            is_checked = bool(row[f'has_SO_{k_key}'])
            data[k_title] = "V" if is_checked else ""
            
        final_rows.append(data)
        
    return pd.DataFrame(final_rows)

# ==========================================
# 4. Excel 輸出
# ==========================================

def export_to_excel(df_all, filename_prefix):
    if df_all.empty:
        print(f"{filename_prefix} 無符合資料。")
        return

    output_filename = f'{filename_prefix}_必修專題課程個別分析_{today_str}.xlsx'
    full_path = os.path.join(OUTPUT_DIR_PATH, output_filename)
    print(f"正在建立 Excel: {full_path}")
    
    # 移除 dept_code
    df_export_base = df_all.drop(columns=['dept_code'])
    
    # 定義欄位順序 (學年插在老師與學期之間)
    base_cols = ['課號', '課程名稱', '開課老師', '學年', '開課學期', '學分數', '必修/選修', '課程類型']
    k_cols = list(k_headers.values())
    stat_cols = ['修課人數', '評量方式', '平均成績', '及格率(%)']
    
    final_col_order = base_cols + k_cols + stat_cols
    
    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
        years = sorted(df_export_base['學年'].unique())
        
        for year in years:
            sheet_name = f"{year}學年度"
            # 這裡不 drop 學年，因為表單內需要顯示
            df_year = df_export_base[df_export_base['學年'] == year].copy()
            
            df_year = df_year[final_col_order]
            
            ws = writer.book.create_sheet(sheet_name)
            rows = dataframe_to_rows(df_year, index=False, header=True)
            
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # 樣式：垂直置中、自動換行 (關鍵)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                         top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    # 對於評量方式欄位 (靠左對齊比較好看)
                    # 評量方式是第 10 欄 (J)
                    if c_idx == 14: # (8 base + 5 K + 1 count + 1 assess) -> 15th column index?
                        # Let's count:
                        # 1:課號, 2:課名, 3:老師, 4:學年, 5:學期, 6:學分, 7:必選, 8:類型
                        # 9~13: K1~K5
                        # 14: 人數
                        # 15: 評量方式
                        # 16: 平均
                        # 17: 及格率
                        if c_idx == 15:
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                    if r_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        ws.row_dimensions[r_idx].height = 60
            
            # 設定欄寬
            ws.column_dimensions['A'].width = 12 # 課號
            ws.column_dimensions['B'].width = 25 # 課名
            ws.column_dimensions['C'].width = 12 # 老師
            ws.column_dimensions['D'].width = 8  # 學年
            ws.column_dimensions['E'].width = 8  # 學期
            ws.column_dimensions['F'].width = 6  # 學分
            ws.column_dimensions['G'].width = 8  # 必選修
            ws.column_dimensions['H'].width = 15 # 類型
            
            # K1-K5 (I~M)
            for i in range(5):
                col_letter = chr(73 + i) # I, J, K, L, M
                ws.column_dimensions[col_letter].width = 15 
            
            ws.column_dimensions['N'].width = 10 # 人數
            ws.column_dimensions['O'].width = 40 # 評量方式 (加寬顯示列表)
            ws.column_dimensions['P'].width = 10 # 平均
            ws.column_dimensions['Q'].width = 10 # 及格率
            
            if 'Sheet' in writer.book.sheetnames:
                writer.book.remove(writer.book['Sheet'])

# ==========================================
# 5. 主程式執行
# ==========================================
try:
    print("開始處理資料...")
    df_full = merge_data()
    
    if not df_full.empty:
        df_full = df_full.sort_values(by=['學年', '開課學期', '課號'])
        
        # 分流
        df_u = df_full[df_full['dept_code'] == 'B301'].copy()
        export_to_excel(df_u, "大學部")
        
        df_g = df_full[df_full['dept_code'] == 'M301'].copy()
        export_to_excel(df_g, "碩士班")
        
        print("-" * 30)
        print(f"全部完成！請檢查: {OUTPUT_DIR_PATH}")
    else:
        print("未找到符合條件的課程。")

except Exception as e:
    print(f"發生錯誤: {e}")
    import traceback
    traceback.print_exc()
finally:
    db.close()