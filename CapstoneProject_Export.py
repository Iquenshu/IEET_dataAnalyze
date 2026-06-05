import os
import pyodbc
import pandas as pd
from datetime import datetime
from openpyxl.styles import Alignment, Border, Side, Font

# ==============================================================================
# 【使用者設定區】 方便您日後修改路徑
# ==============================================================================
# 1. 目的地 Excel 輸出資料夾路徑
OUTPUT_DIR = r"D:\113年後資料\系辦辦公相關\IEET認證\python程式\PythonIEET\PythonIEET\output_files\專題成員名單"

# 2. 您的 IEET 來源資料庫路徑
TARGET_DB = r"D:\113年後資料\系辦辦公相關\IEET認證\python程式\PythonIEET\PythonIEET\IEETdatabase.accdb"

# 3. 資料來源資料表名稱
SRC_TABLE_NAME = "CapstoneProject_Data"
# ==============================================================================

# Access 連線字串範本
CONN_STR = r"Driver={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={};".format(TARGET_DB)

def excel_formatting(ws, num_rows):
    """
    對 Excel 進行格式美化，使其外觀完全符合使用者提供的附圖樣式
    """
    # 定義基本樣式 (使用台灣常用的微軟正黑體)
    header_font = Font(name='微軟正黑體', size=12, bold=True)
    data_font = Font(name='微軟正黑體', size=11)
    
    # 對齊方式
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 全黑細框線
    thin_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # 1. 設定標題列樣式 (第 1 列)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # 2. 設定資料列樣式 (第 2 列到最後一列)
    for row in range(2, num_rows + 2):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            
            # 根據欄位性質設定對齊：學年度(1)、團隊編號(2)、指導教授(5) 置中；其餘靠左
            if col in [1, 2, 5]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # 3. 將第一欄的「學年度」進行垂直合併 (從第 2 列合併到最後一列)
    if num_rows > 1:
        ws.merge_cells(start_row=2, start_column=1, end_row=num_rows + 1, end_column=1)
        # 合併後需重新將左上角的主儲存格置中
        ws.cell(row=2, column=1).alignment = center_align

    # 4. 固定各欄位的舒適寬度 (已調大 B 欄寬度以完整顯示專題組編號)
    ws.column_dimensions['A'].width = 14  # 學年度
    ws.column_dimensions['B'].width = 20  # 團隊編號 (顯示原始專題組編號，利於校對)
    ws.column_dimensions['C'].width = 25  # 團隊成員名單
    ws.column_dimensions['D'].width = 40  # 專題名稱
    ws.column_dimensions['E'].width = 15  # 指導教授


def main():
    # 檢查並自動建立輸出資料夾
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        print(f"已自動建立輸出資料夾：{OUTPUT_DIR}")

    if not os.path.exists(TARGET_DB):
        print(f"錯誤：找不到資料庫檔案 {TARGET_DB}")
        return

    print("開始讀取資料庫...")
    # 步驟一：連線資料庫並將整張表載入成 Pandas DataFrame
    conn = pyodbc.connect(CONN_STR)
    query = f"SELECT * FROM {SRC_TABLE_NAME}"
    df = pd.read_sql(query, conn)
    conn.close()

    if df.empty:
        print("資料庫內沒有任何資料，取消匯出。")
        return

    # 清理資料字串前後的空白，並處理 Null 值
    df = df.fillna("").astype(str).apply(lambda x: x.str.strip())
    df = df.replace('None', '')

    print("正在進行團隊成員名單整合...")
    # 步驟二：核心群組邏輯
    # 依 學年度、參賽分組、專題組編號 進行分組，將同組學生的姓名用「、」串接起來
    df_grouped = df.groupby(['學年度', '參賽分組', '專題組編號']).agg({
        '學生姓名': lambda x: "、".join([name for name in x.unique() if name]), # 串接學生成員
        '專題名稱': 'first', 
        '指導教授': 'first'  
    }).reset_index()

    # 重新命名欄位名稱
    df_grouped = df_grouped.rename(columns={'學生姓名': '團隊成員名單'})

    # 獲取今天的日期 (格式如: 20260604)
    today_str = datetime.now().strftime("%Y%m%d")

    # 步驟三：依「大學部」與「研究所」分流
    classifications = ['大學部', '研究所']

    for cls in classifications:
        # 篩選出該分組的資料
        df_cls = df_grouped[df_grouped['參賽分組'] == cls].copy()
        
        if df_cls.empty:
            print(f"找不到任何【{cls}】的資料，跳過該檔案製作。")
            continue

        # 設定該分組的 Excel 完整輸出路徑
        file_name = f"{cls}_專題成員名單_{today_str}.xlsx"
        excel_path = os.path.join(OUTPUT_DIR, file_name)

        print(f"\n正在建構【{cls}】的 Excel 檔案...")
        
        # 取得該分組內所有的學年度，並由小到大排序
        available_years = sorted(df_cls['學年度'].unique())

        # 開始寫入 Excel
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for year in available_years:
                # 篩選特定學年度的資料
                df_year = df_cls[df_cls['學年度'] == year].copy()
                
                # 排序專題組編號
                df_year = df_year.sort_values(by='專題組編號')

                # 【修改處】直接將原始的「專題組編號」指定給「團隊編號」欄位，方便您校對
                df_year['團隊編號'] = df_year['專題組編號']

                # 挑選最終要呈現的 5 個欄位並排出順序
                df_final_output = df_year[['學年度', '團隊編號', '團隊成員名單', '專題名稱', '指導教授']]

                # 將資料寫入該學年度的分頁 (Sheet)
                sheet_name = str(year) if year else "未分類學年度"
                df_final_output.to_excel(writer, sheet_name=sheet_name, index=False)

                # 取得當前工作表物件，進行畫線與垂直合併等美化
                ws = writer.sheets[sheet_name]
                excel_formatting(ws, len(df_final_output))

        print(f"成功產出檔案：{file_name}")

    print("\n==============================================")
    print("所有 IEET 專題 Excel 報表（含原始編號）已成功產出！")
    print("==============================================")

if __name__ == "__main__":
    main()