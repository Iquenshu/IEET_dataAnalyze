import pandas as pd
import os
from Accessdb import AccessHelper
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

# ==========================================
# 1. 全域路徑與環境設定
# ==========================================
db = AccessHelper()

# 輸出目錄與檔名設定 (自動帶入當前執行日期)
OUTPUT_DIR = r'output_files\課程分類'
TODAY_STR = pd.Timestamp.now().strftime("%Y%m%d")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'課程分類表_由資料庫生成_{TODAY_STR}.xlsx')

# 確保輸出資料夾存在
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)
    print(f"建立目錄: {OUTPUT_DIR}")


# ==========================================
# 2. 核心層級式互斥分類邏輯 (嚴格遵循優先權規範)
# ==========================================
def classify_course_strict(row):
    """
    依據最嚴格的優先權層級進行互斥分類 (4選1)。
    程式由上至下執行，前方的關卡具有絕對攔截權，確保不違反優先規則。
    """
    name = str(row['課程名稱']).strip()
    name_upper = name.upper()
    
    # 初始化所有分類註記為 0
    is_math = 0
    is_science = 0
    is_eng = 0
    is_general = 0
    
    # ---------------------------------------------------------
    # 🌟 關卡 0：計算機核心特例攔截點 (最高優先)
    # ---------------------------------------------------------
    if '計算機' in name_upper:
        is_science = 1
        return is_math, is_science, is_eng, is_general

    # ---------------------------------------------------------
    # 🌟 關卡 1：嚴格 Priority 1 - 數學 (Mathematics)
    # ---------------------------------------------------------
    math_keywords = [
        # 優先規則 1 指定科目
        '微積分', '微分方程', '線性代數', '機率與統計', '機率', '統計',
        # 系統分流特例
        '線性系統', '非線性系統', 
        # 課名有「數學」就丟到數學
        '數學', '應用數學', '數學導論', '模糊數學',
        # 歷年既有核心數學與理論字根
        '複變函數', '複變', '數值分析', '迴歸分析', '變量分析', '模糊理論', 
        '資料壓縮', '矩陣理論', '數值運算', '希爾伯特空間', '軟性計算', '賽局', 
        '最佳化', '隨機過程', '圖論', '圖形理論', '資料探勘', '數據分析', 
        '資料分析', '大數據', '消息理論', '群論', '計算理論', '智慧資料處理', 
        '商務數據', '圖型識別', '圖訊識別', '離散', '幾何', '代數', '計算', '數值', '運算',
        'CALCULUS', 'ALGEBRA', 'PROBABILITY', 'STATISTICS', 'MATH'
    ]
    
    # ---------------------------------------------------------
    # 🌟 關卡 2：嚴格 Priority 2 - 基礎科學 (Science)
    # ---------------------------------------------------------
    sci_keywords = [
        # 優先規則 2 指定科目
        '普通物理', 
        # 系統分流特例
        '訊號與系統', '數位系統設計',
        # 📌 新增指定科目：生理學、生理、圖學
        '生理學', '生理', '圖學', 
        # 既有核心科學、生物、物理理論字根
        '工程聲學', '聲學', '材料科學', '藻類', '生藥', '測量學', '科學思考',
        '凝態理論', '電磁理論', '物理', '化學', '生物', '力學', '熱力學', '量子', '光學', 
        'PHYSICS', 'CHEMISTRY', 'BIOLOGY', 'MECHANICS'
    ]

    # ---------------------------------------------------------
    # 🌟 關卡 3：嚴格 Priority 3 - 工程專業 (Engineering Professional)
    # ---------------------------------------------------------
    eng_keywords = [
        # 優先規則 3 指定科目
        '電路學', '電子學', '電磁學', '電機機械', '通訊系統', '控制系統', 
        '電工實驗', '電機工程進階實作專案', '實作專題', '書報討論',
        # 📌 新增與優化通用專業字根：語音、儀器、生醫儀器、化合物半導體、語音處理
        '生醫儀器', '化合物半導體', '語音處理', '語音', '儀器', 
        # 通用丟工程字根
        '載具', '無人載具', '風電', '離岸風電', '通信', '程式', '電磁', '通訊', '信號', '材料', '深度學習',
        # 系統與硬體控制通用字根
        '系統', '元件', '控制', '工程', '電子', '作業系統', '微波', '技術', '傳輸', '半導體',
        # 其他既有專業工程領域延伸字根
        '內燃機', '電磁波', '生醫影像', '磁共振影像', '磁振造影', '腦機介面', '人工智慧', 
        '能量轉換', '配電', '工業配電', '保護電驛', '機械製造', '氫能', '燃料電池', 
        '智慧製造', '智慧製造聯網', '金融科技', '高分子', '機器人', '矽覆絕緣', 
        '感測', '奈米', '電腦模擬', '電腦視覺', '三維視覺', '深度視覺', '影像辨識', 
        '機器視覺', '影像處理', '駭客攻防', '電腦鑑識', '陶瓷元件', '空間調查', 
        '微影', '薄膜技術', '表面分析', '可靠度', '微系統', '奈微系統', '嵌入式',
        '智慧型控制', '最佳控制', '生醫工程', '電子材料', '工程材料', '電激發光', '積體電路',
        '微機電', '機電材料', '記憶體', '太陽能', '錯誤更正', '3D內容', '3D列印', 
        '海域調查', '研究方法', '視訊編碼', '光纖', 'SMALL CELL', '聯網整合', 
        '實務專案', '輔助繪圖', '技術繪圖', '繪圖', '電機', '資訊', '晶片', 
        '電力', '多媒體', '設計', '實習', '專題', '實驗', '邏輯', '微處理', 
        'VLSI', 'FPGA', 'JAVA', 'PYTHON', 'C++', 'AI', '機器學習', '演算法', 
        '資料結構', '網路', '訊號', '電波', '光電', '類比', '數位', '軟體', '程式語言', '硬體描述語言', 'HDL', '計算機',
        'ELECTRIC', 'ELECTRONIC', 'SYSTEM', 'SIGNAL', 'CONTROL', 'COMMUNICATION',
        'NETWORK', 'SEMICONDUCTOR', 'CHIP', 'DESIGN', 'PROJECT', 'LAB'
    ]

    # --- 🛠️ 嚴格層級式分流判斷 ---
    if any(k in name_upper for k in math_keywords):
        is_math = 1
    elif any(k in name_upper for k in sci_keywords):
        is_science = 1
    elif any(k in name_upper for k in eng_keywords):
        is_eng = 1
    else:
        # 完美收容：凡是不屬於 數學、科學、工程 專業領域的，全數自動歸檔為通識
        is_general = 1

    return is_math, is_science, is_eng, is_general


# ==========================================
# 3. 主程式：由資料庫生成清單
# ==========================================
def generate_classification_from_db():
    # A. 改由 course_code (課號) 欄位建立本系教師比對字典
    print("正在從資料庫 Courses 資料表建立本系【課號 ➔ 教師】比對字典...")
    instructor_map = {}
    try:
        df_ee_courses = pd.read_sql("SELECT course_code, instructor FROM Courses", db.conn)
        for _, row_ee in df_ee_courses.iterrows():
            if pd.notna(row_ee['course_code']):
                ee_code = str(row_ee['course_code']).strip()
                ee_inst = str(row_ee['instructor']).strip() if pd.notna(row_ee['instructor']) else ""
                if ee_inst:
                    instructor_map[ee_code] = ee_inst
        print(f"➔ 成功載入 {len(instructor_map)} 筆本系核心課號之教師對應資料。")
    except Exception as e:
        print(f"⚠️ 讀取 Courses 表失敗，教師欄位將改為全非本系模式: {e}")

    # B. 從 STscore 讀取全體學生不重複的修課清單
    print("正在從資料庫 STscore 資料表讀取歷史修課清單...")
    sql = "SELECT DISTINCT 學年度, 學期, 課號, 課程名稱, 學分數 FROM STscore"
    try:
        df_courses = pd.read_sql(sql, db.conn)
    except Exception as e:
        print(f"❌ 資料庫讀取失敗: {e}")
        return

    if df_courses.empty:
        print("⚠️ 警告：STscore 資料表無任何數據！")
        return

    print(f"成功取得 {len(df_courses)} 筆開課歷史紀錄。開始執行跨表【課號】教師比對與層級分流...")

    results = []
    for _, row in df_courses.iterrows():
        c_code = str(row['課號']).strip()
        c_name = str(row['課程名稱']).strip()
        
        # 執行層級優先權自動分類
        m, s, e, g = classify_course_strict(row)
        
        # 授課教師比對：改以 c_code (課號) 向字典進行提領
        inst_name = instructor_map.get(c_code, "非本系開設")
        if not inst_name or inst_name.lower() == 'nan':
            inst_name = "非本系開設"
        
        # 將資料依指定順序塞入
        results.append({
            '學年度': row['學年度'],
            '學期': row['學期'],
            '課號': c_code,
            '課程名稱': c_name,
            '學分': row['學分數'],
            '授課教師': inst_name,  
            'is_math': m,
            'is_science': s,
            'is_eng_prof': e,
            'is_general': g
        })
        
    df_out = pd.DataFrame(results)
    
    # 報表排序：將所有「通識課程 (is_general == 1)」全面高亮置頂，其餘依學年度、學期、課號排列
    df_out = df_out.sort_values(
        by=['is_general', '學年度', '學期', '課號'], 
        ascending=[False, False, True, True]
    )
    
    print(f"正在將分類結果寫入 Excel 報表: {OUTPUT_FILE}")
    
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df_out.to_excel(writer, index=False, sheet_name='課程分類')
        ws = writer.sheets['課程分類']
        
        # 套用高質感淺綠色填充樣式 (全面標記所有通識課程)
        general_fill = PatternFill(start_color="E2EFDA", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 標頭美化 (A-J 欄，共 10 大欄位)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # 巡覽資料：格式設定與通識顏色填充
        general_count = 0
        for i, row_data in enumerate(df_out.itertuples(), start=2):
            is_gen = (row_data.is_general == 1)
            if is_gen:
                general_count += 1
                
            for col in range(1, 11): # A 到 J 欄
                cell = ws.cell(row=i, column=col)
                cell.border = thin_border
                
                # 所有通識課程整行塗上淺綠色
                if is_gen:
                    cell.fill = general_fill
                
                # 排版控制：除課程名稱(第4欄)靠左外，其餘欄位一律全面置中
                if col != 4:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
        # 自動適應欄寬
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 11)

    print("==================================================")
    print(f"🎉 歷年課程分類清單（含新專業選修優化）建置完成！")
    print(f"總處理開課歷史紀錄: {len(df_out)} 筆")
    print(f"🟢 劃分至通識類之課程: {general_count} 筆 (已全面高亮淺綠色並置頂)")
    print("==================================================")

if __name__ == "__main__":
    generate_classification_from_db()
    db.close()