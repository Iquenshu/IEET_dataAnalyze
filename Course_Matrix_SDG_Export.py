import pandas as pd
from Accessdb import AccessHelper
import os
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import warnings

# 忽略 Pandas 與 SQLAlchemy 的警告訊息
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)

# ==========================================
# 1. 設定與準備
# ==========================================
db = AccessHelper()
today_str = datetime.today().strftime('%Y%m%d')

BASE_DIR = 'output_files'
SUB_DIR = '課程核心能力及SDG關聯表'
OUTPUT_DIR_PATH = os.path.join(BASE_DIR, SUB_DIR)

if not os.path.exists(OUTPUT_DIR_PATH):
    os.makedirs(OUTPUT_DIR_PATH)
    print(f"已建立資料夾: {OUTPUT_DIR_PATH}")

k_mapping = {
    'has_SO_K1': 'K1 能夠整合、組織電機專業理論來分析、表達問題之能力',
    'has_SO_K2': 'K2 能夠運用電機專業知識解決及實作電機工程問題之能力',
    'has_SO_K3': 'K3 具備分工、協調、重視團隊合作精神、遵守工程倫理以達成工作目標之能力',
    'has_SO_K4': 'K4 能夠激發自己潛能、融合他人智慧，具備獨立思考以及研究創新之能力',
    'has_SO_K5': 'K5 具備吸收電機新知、掌握國際發展趨勢，隨時接受競爭挑戰之能力'
}

sdg_mapping = {
    'sdg_1': 'SDG1 消除貧窮',
    'sdg_2': 'SDG2 消除飢餓',
    'sdg_3': 'SDG3 良好健康與福祉',
    'sdg_4': 'SDG4 教育品質',
    'sdg_5': 'SDG5 性別平等',
    'sdg_6': 'SDG6 乾淨水源與公共衛生',
    'sdg_7': 'SDG7 可負擔乾淨能源',
    'sdg_8': 'SDG8 優質工作與經濟成長',
    'sdg_9': 'SDG9 工業、創新和基礎建設',
    'sdg_10': 'SDG10 減少不平等',
    'sdg_11': 'SDG11 永續城市',
    'sdg_12': 'SDG12 責任消費與生產',
    'sdg_13': 'SDG13 氣候行動',
    'sdg_14': 'SDG14 海洋生態',
    'sdg_15': 'SDG15 陸域生態',
    'sdg_16': 'SDG16 和平、正義和穩健的制度',
    'sdg_17': 'SDG17 促進目標實現的全球夥伴關係'
}

k_cols = list(k_mapping.keys())
sdg_cols = list(sdg_mapping.keys())

# ==========================================
# 2. 資料讀取
# ==========================================
print("正在讀取並整合資料庫 (Courses + Matrix + SDGs)...")

sql = '''
SELECT 
    C.academic_year, C.semester, C.dept_code,
    C.course_code, C.course_name, C.is_required, C.credits, C.instructor,
    M.has_SO_K1, M.has_SO_K2, M.has_SO_K3, M.has_SO_K4, M.has_SO_K5,
    S.sdg_1, S.sdg_2, S.sdg_3, S.sdg_4, S.sdg_5, S.sdg_6, 
    S.sdg_7, S.sdg_8, S.sdg_9, S.sdg_10, S.sdg_11, S.sdg_12, 
    S.sdg_13, S.sdg_14, S.sdg_15, S.sdg_16, S.sdg_17
FROM (Courses AS C
LEFT JOIN Course_Matrix AS M ON C.id = M.course_id)
LEFT JOIN Course_SDGs AS S ON C.id = S.course_id
ORDER BY C.academic_year DESC, C.semester, C.course_code
'''

df_raw = pd.read_sql(sql, db.conn)
df_raw = df_raw[df_raw['semester'] != 3]

def format_check_numeric(val):
    if pd.isna(val): return 0
    try:
        if isinstance(val, (bool, int, float)):
            return 1 if int(val) == 1 else 0
        s_val = str(val).lower().strip()
        if s_val in ['1', 'true', '1.0', 'yes']: return 1
        return 0
    except: return 0

def get_required_str(val):
    try:
        if isinstance(val, bool): return "必修" if val else "選修"
        s_val = str(val).lower().strip()
        if s_val in ['1', 'true', '1.0', 'yes']: return "必修"
        return "選修"
    except: return "選修"

def process_data_for_table(df_sem, target_cols, mapping_dict):
    output_rows = []
    display_labels = [mapping_dict[c] for c in target_cols]
    totals = {label: 0 for label in display_labels}
    
    for _, row in df_sem.iterrows():
        row_sum = 0
        current_vals = []
        for col_db in target_cols:
            val = format_check_numeric(row[col_db])
            row_sum += val
            current_vals.append(val)
            
        if row_sum == 0: continue

        excel_row = {
            '學年': row['academic_year'],
            '學期': row['semester'],
            '課程名稱及課號': f"{row['course_name']} ({row['course_code']})",
            '必選修': get_required_str(row['is_required']),
            '學分': row['credits']
        }
        
        for val, col_label in zip(current_vals, display_labels):
            excel_row[col_label] = val
            totals[col_label] += val
                
        output_rows.append(excel_row)
        
    df_out = pd.DataFrame(output_rows)
    base_cols = ['學年', '學期', '課程名稱及課號', '必選修', '學分']
    if not df_out.empty:
        df_out = df_out[base_cols + display_labels]
    else:
        df_out = pd.DataFrame(columns=base_cols + display_labels)
    
    return df_out, totals, display_labels

def get_missing_string(df_sem, check_cols):
    missing_items = []
    for _, row in df_sem.iterrows():
        row_sum = 0
        for col in check_cols:
            row_sum += format_check_numeric(row[col])
        if row_sum == 0:
            missing_items.append(f"{row['course_name']} ({row['course_code']})")
    return ", ".join(missing_items) if missing_items else "(無)"

def write_table_block(ws, df_data, totals, labels, start_row, header_color):
    rows = list(dataframe_to_rows(df_data, index=False, header=True))
    for r_idx, row in enumerate(rows):
        current_excel_row = start_row + r_idx
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=current_excel_row, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            if r_idx == 0:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                ws.row_dimensions[current_excel_row].height = 60
    
    last_row = start_row + len(df_data) + 1 
    ws.cell(row=last_row, column=1, value="總計").font = Font(bold=True)
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=5)
    
    cell_total_title = ws.cell(row=last_row, column=1)
    cell_total_title.alignment = Alignment(horizontal='center')
    cell_total_title.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    cell_total_title.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    start_col_idx = 6
    for i, label in enumerate(labels):
        col_idx = start_col_idx + i
        cell = ws.cell(row=last_row, column=col_idx, value=totals[label])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    return last_row + 4

def write_missing_list_section(ws, missing_k_str, missing_sdg_str, start_row):
    ws.cell(row=start_row, column=1, value="未關聯任何核心能力之課程").font = Font(bold=True)
    ws.cell(row=start_row, column=1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    cell_content_k = ws.cell(row=start_row, column=2, value=missing_k_str)
    cell_content_k.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=15)
    ws.row_dimensions[start_row].height = 30
    
    next_row = start_row + 1
    ws.cell(row=next_row, column=1, value="未關聯任何 SDGs 之課程").font = Font(bold=True)
    ws.cell(row=next_row, column=1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    cell_content_sdg = ws.cell(row=next_row, column=2, value=missing_sdg_str)
    cell_content_sdg.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=15)
    ws.row_dimensions[next_row].height = 30
    return next_row + 2

def set_column_widths(ws, label_count):
    ws.column_dimensions['A'].width = 8  
    ws.column_dimensions['B'].width = 6  
    ws.column_dimensions['C'].width = 40 
    ws.column_dimensions['D'].width = 10 
    ws.column_dimensions['E'].width = 6  
    
    for i in range(label_count):
        col_letter = chr(64 + 6 + i)
        if 6 + i > 26:
            col_letter = chr(64 + (6 + i - 1) // 26) + chr(64 + (6 + i - 1) % 26 + 1)
        ws.column_dimensions[col_letter].width = 15

# ==========================================
# 建立 K1~K5 與 SDG1~SDG17 兩張歷年統計摘要表
# ==========================================
def create_summary_dashboard(ws, df_dept, dept_name):
    ws.views.sheetView[0].showGridLines = True
    unique_years = sorted(df_dept['academic_year'].unique(), reverse=False)
    
    # ----------------------------------------------------------------------
    # 表格一：課程核心能力 (K1~K5) 歷年統計摘要
    # ----------------------------------------------------------------------
    ws.cell(row=2, column=2, value=f"{dept_name} 歷年核心能力符合課程數統計摘要").font = Font(size=13, bold=True)
    
    headers_k = ["統計學年 / 項目"] + [k_mapping[k] for k in k_cols]
    for c_idx, h_text in enumerate(headers_k, 2):
        cell = ws.cell(row=4, column=c_idx, value=h_text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid") # 橘色
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.row_dimensions[4].height = 55
    
    current_row = 5
    for year in unique_years:
        df_year = df_dept[df_dept['academic_year'] == year]
        cell_yr = ws.cell(row=current_row, column=2, value=f"{int(year)}學年度")
        cell_yr.alignment = Alignment(horizontal='center')
        cell_yr.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for c_idx, k_col in enumerate(k_cols, 3):
            count = sum(df_year[k_col].apply(format_check_numeric))
            cell = ws.cell(row=current_row, column=c_idx, value=count)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        current_row += 1
        
    # K對比表總計列
    cell_tot_lbl1 = ws.cell(row=current_row, column=2, value="歷年全部總和")
    cell_tot_lbl1.font = Font(bold=True)
    cell_tot_lbl1.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    cell_tot_lbl1.alignment = Alignment(horizontal='center')
    cell_tot_lbl1.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for c_idx, k_col in enumerate(k_cols, 3):
        total_count = sum(df_dept[k_col].apply(format_check_numeric))
        cell = ws.cell(row=current_row, column=c_idx, value=total_count)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
    # ----------------------------------------------------------------------
    # 表格二：SDGs (SDG1~SDG17) 歷年統計摘要 (新加入)
    # ----------------------------------------------------------------------
    start_sdg_section = current_row + 4
    ws.cell(row=start_sdg_section - 1, column=2, value=f"{dept_name} 歷年 SDGs 符合課程數統計摘要").font = Font(size=13, bold=True)
    
    headers_sdg = ["統計學年 / 項目"] + [sdg_mapping[s] for s in sdg_cols]
    for c_idx, h_text in enumerate(headers_sdg, 2):
        cell = ws.cell(row=start_sdg_section, column=c_idx, value=h_text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid") # 綠色
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.row_dimensions[start_sdg_section].height = 55
    
    current_row = start_sdg_section + 1
    for year in unique_years:
        df_year = df_dept[df_dept['academic_year'] == year]
        cell_yr = ws.cell(row=current_row, column=2, value=f"{int(year)}學年度")
        cell_yr.alignment = Alignment(horizontal='center')
        cell_yr.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for c_idx, sdg_col in enumerate(sdg_cols, 3):
            count = sum(df_year[sdg_col].apply(format_check_numeric))
            cell = ws.cell(row=current_row, column=c_idx, value=count)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        current_row += 1
        
    # SDG對比表總計列
    cell_tot_lbl2 = ws.cell(row=current_row, column=2, value="歷年全部總和")
    cell_tot_lbl2.font = Font(bold=True)
    cell_tot_lbl2.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    cell_tot_lbl2.alignment = Alignment(horizontal='center')
    cell_tot_lbl2.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for c_idx, sdg_col in enumerate(sdg_cols, 3):
        total_count = sum(df_dept[sdg_col].apply(format_check_numeric))
        cell = ws.cell(row=current_row, column=c_idx, value=total_count)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 調整特定 A、B 欄與動態數據欄寬度，防止文字截斷
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 18
    for i in range(len(sdg_cols)):
        col_letter = chr(67 + i)
        if 3 + i > 26: # 超過 Z 欄處理
            col_letter = chr(64 + (3 + i - 1) // 26) + chr(64 + (3 + i - 1) % 26 + 1)
        ws.column_dimensions[col_letter].width = 24

# ==========================================
# 6. 主流程
# ==========================================
def process_dept_export(df_dept, dept_name):
    if df_dept.empty:
        print(f"{dept_name} 無資料。")
        return

    output_filename = f'{dept_name}_課程核心能力及SDG關聯表_{today_str}.xlsx'
    full_path = os.path.join(OUTPUT_DIR_PATH, output_filename)
    
    print(f"正在建立檔案: {full_path}")
    
    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
        
        # 1. 建立高精簡、無圖表的「歷年全部總計」統計分頁 (含核心能力表 與 SDG 統計表)
        ws_all = writer.book.create_sheet("歷年全部總計")
        create_summary_dashboard(ws_all, df_dept, dept_name)
        
        # 2. 依序建立各學年與學期細節分頁
        unique_years = sorted(df_dept['academic_year'].unique(), reverse=True)
        
        for year in unique_years:
            year = int(year)
            
            df_year_total = df_dept[df_dept['academic_year'] == year].copy()
            df_year_total = df_year_total.sort_values(
                by=['is_required', 'semester', 'course_code'], 
                ascending=[False, True, True]
            )
            
            ws_year = writer.book.create_sheet(f"{year}全學年總計")
            ws_year.views.sheetView[0].showGridLines = True
            current_row_y = 1
            
            df_k_y, totals_k_y, labels_k_y = process_data_for_table(df_year_total, k_cols, k_mapping)
            current_row_y = write_table_block(ws_year, df_k_y, totals_k_y, labels_k_y, current_row_y, "ED7D31")
            
            df_sdg_y, totals_sdg_y, labels_sdg_y = process_data_for_table(df_year_total, sdg_cols, sdg_mapping)
            current_row_y = write_table_block(ws_year, df_sdg_y, totals_sdg_y, labels_sdg_y, current_row_y, "70AD47")
            
            missing_k_str_y = get_missing_string(df_year_total, k_cols)
            missing_sdg_str_y = get_missing_string(df_year_total, sdg_cols)
            current_row_y += 1
            write_missing_list_section(ws_year, missing_k_str_y, missing_sdg_str_y, current_row_y)
            set_column_widths(ws_year, 17)
            
            for sem in [1, 2]:
                df_sem = df_year_total[df_year_total['semester'] == sem].copy()
                if df_sem.empty: continue
                    
                sheet_name = f"{year}-{sem}"
                df_sem = df_sem.sort_values(by=['is_required', 'course_code'], ascending=[False, True])
                
                ws_sem = writer.book.create_sheet(sheet_name)
                ws_sem.views.sheetView[0].showGridLines = True
                current_row_s = 1
                
                df_k_s, totals_k_s, labels_k_s = process_data_for_table(df_sem, k_cols, k_mapping)
                current_row_s = write_table_block(ws_sem, df_k_s, totals_k_s, labels_k_s, current_row_s, "ED7D31")
                
                df_sdg_s, totals_sdg_s, labels_sdg_s = process_data_for_table(df_sem, sdg_cols, sdg_mapping)
                current_row_s = write_table_block(ws_sem, df_sdg_s, totals_sdg_s, labels_sdg_s, current_row_s, "70AD47")
                
                missing_k_str_s = get_missing_string(df_sem, k_cols)
                missing_sdg_s_str = get_missing_string(df_sem, sdg_cols)
                current_row_s += 1
                write_missing_list_section(ws_sem, missing_k_str_s, missing_sdg_s_str, current_row_s)
                set_column_widths(ws_sem, 17)
                
        if 'Sheet' in writer.book.sheetnames:
            writer.book.remove(writer.book['Sheet'])
            
    print(f"匯出成功！")

# 執行
if not df_raw.empty:
    df_raw['dept_code'] = df_raw['dept_code'].astype(str).str.strip()
    
    # 大學部
    df_undergrad = df_raw[df_raw['dept_code'] == 'B301'].copy()
    process_dept_export(df_undergrad, "大學部")
    
    # 碩士班
    df_grad = df_raw[df_raw['dept_code'] == 'M301'].copy()
    process_dept_export(df_grad, "碩士班")
else:
    print("資料庫無任何課程資料。")

db.close()
print("-" * 30)
print(f"全部完成！請檢查: {OUTPUT_DIR_PATH}")