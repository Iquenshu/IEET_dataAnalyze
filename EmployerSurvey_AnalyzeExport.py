import pandas as pd
from Accessdb import AccessHelper
import os
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ==========================================
# 1. 設定與準備
# ==========================================
db = AccessHelper()
today_str = datetime.today().strftime('%Y%m%d')

# [設定] 檔案輸出路徑
BASE_DIR = 'output_files'
SUB_DIR = '雇主問券分析'
OUTPUT_DIR_PATH = os.path.join(BASE_DIR, SUB_DIR)

# 自動建立資料夾
if not os.path.exists(OUTPUT_DIR_PATH):
    os.makedirs(OUTPUT_DIR_PATH)
    print(f"已建立資料夾: {OUTPUT_DIR_PATH}")

# 定義核心能力對應
competency_mapping = {
    'K1': 'Q7_Theory_Perf',
    'K2': 'Q8_Tech_Perf',
    'K3': 'Q9_Team_Perf',
    'K4': 'Q10_Innov_Perf',
    'K5': 'Q11_Global_Perf'
}

core_competencies = {
    'K1': '能夠整合、組織電機專業理論來分析、表達問題之能力。',
    'K2': '能夠運用電機專業知識解決及實作電機工程問題之能力。',
    'K3': '具備分工、協調、重視團隊合作精神、遵守工程倫理以達成工作目標之能力。',
    'K4': '能夠激發自己潛能、融合他人智慧，具備獨立思考以及研究創新之能力。',
    'K5': '具備吸收電機新知、掌握國際發展趨勢，隨時接受競爭挑戰之能力。'
}

# 分數轉換表 (根據您的要求)
# 為了程式健壯性，我們會做部分關鍵字比對
score_mapping = {
    '非常不滿意': 20,
    '不太滿意': 40,
    '普通': 60,
    '滿意': 80,       # 注意：這可能會跟 '非常滿意' 衝突，需小心比對順序或完整比對
    '非常滿意': 100
}

def parse_score(val):
    """
    將問卷文字轉換為分數
    """
    if pd.isna(val):
        return None
    
    s = str(val).strip()
    
    # 優先比對完整字串 (如果您資料庫裡的字串跟提供的一模一樣)
    if '非常滿意' in s: return 100
    if '不太滿意' in s: return 40
    if '非常不滿意' in s: return 20
    if '普通' in s: return 60
    if '滿意' in s: return 80 # 最後再比對 "滿意"，以免誤判 "非常滿意" 或 "不太滿意"
    
    return None

def get_academic_year(date_val):
    """
    根據填表日期判斷學年度
    規則：8月1日以後為新學年
    例如：2023/09/01 -> 112學年, 2024/02/01 -> 112學年, 2024/08/01 -> 113學年
    """
    if pd.isna(date_val):
        return '未知年份'
    
    try:
        dt = pd.to_datetime(date_val)
        year = dt.year - 1911 # 轉民國年
        month = dt.month
        
        if month >= 8:
            return str(year)
        else:
            return str(year - 1)
    except:
        return '未知年份'

# ==========================================
# 2. 資料處理
# ==========================================
print("正在讀取 EmployerSurvey 資料表...")
df_raw = pd.read_sql("SELECT * FROM EmployerSurvey", db.conn)

if df_raw.empty:
    print("錯誤：EmployerSurvey 資料表是空的。")
    db.close()
    exit()

print(f"取得 {len(df_raw)} 筆原始問卷資料。")

# 1. 計算學年度
df_raw['AcademicYear'] = df_raw['Fill_Date'].apply(get_academic_year)

# 2. 轉換分數
for k, col in competency_mapping.items():
    # 新增欄位例如 K1_Score
    df_raw[f'{k}_Score'] = df_raw[col].apply(parse_score)

# 3. 統計分析
# 依學年分組
yearly_stats = []
years = sorted(df_raw['AcademicYear'].unique())

for year in years:
    if year == '未知年份': continue
    
    df_year = df_raw[df_raw['AcademicYear'] == year]
    count = len(df_year)
    
    row = {'學年': year, '樣本數': count}
    
    for k in ['K1', 'K2', 'K3', 'K4', 'K5']:
        col_score = f'{k}_Score'
        avg = df_year[col_score].mean()
        # 如果全都是 NaN (無有效回答)，則給 0
        if pd.isna(avg): avg = 0.0
        row[k] = round(avg, 2)
        
    yearly_stats.append(row)

df_trend = pd.DataFrame(yearly_stats)

# ==========================================
# 3. Excel 輸出
# ==========================================
output_filename = f'雇主問券核心能力分析_{today_str}.xlsx'
full_path = os.path.join(OUTPUT_DIR_PATH, output_filename)
print(f"正在輸出: {full_path}")

def write_table(ws, title, df_data, start_row=1):
    # 標題
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # 欄位
    headers = ['學年', '有效樣本數'] + [f"{k} {desc}" for k, desc in core_competencies.items()]
    
    current_row = start_row + 1
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=i, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        col_letter = chr(64 + i)
        if i <= 2:
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = 25

    current_row += 1
    
    if df_data.empty:
        ws.cell(row=current_row, column=1, value="(無資料)")
        return
        
    for _, row in df_data.iterrows():
        # 學年
        ws.cell(row=current_row, column=1, value=f"{row['學年']}學年度").alignment = Alignment(horizontal='center')
        # 樣本數
        ws.cell(row=current_row, column=2, value=row['樣本數']).alignment = Alignment(horizontal='center')
        
        # K1-K5
        for i, k in enumerate(['K1', 'K2', 'K3', 'K4', 'K5'], 3):
            val = row.get(k, 0.0)
            cell = ws.cell(row=current_row, column=i, value=val)
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0.00'
            
        current_row += 1

with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
    ws = writer.book.create_sheet("雇主滿意度趨勢分析", 0)
    write_table(ws, "【歷年雇主對畢業生核心能力滿意度分析】", df_trend)
    
    # 如果需要看原始資料，可以把 df_raw 也輸出到另一個分頁
    # ws_raw = writer.book.create_sheet("原始資料清單")
    # ... (略)

    if 'Sheet' in writer.book.sheetnames:
        writer.book.remove(writer.book['Sheet'])

db.close()
print("-" * 30)
print("執行完成！")