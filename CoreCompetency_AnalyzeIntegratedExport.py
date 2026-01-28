import pandas as pd
from Accessdb import AccessHelper
import os
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ==========================================
# 1. 設定與準備
# ==========================================
db = AccessHelper()
today_str = datetime.today().strftime('%Y%m%d')

# [設定] 檔案輸出路徑
BASE_DIR = 'output_files'
SUB_DIR = '核心能力綜合分析'
OUTPUT_DIR_PATH = os.path.join(BASE_DIR, SUB_DIR)

# 自動建立資料夾
if not os.path.exists(OUTPUT_DIR_PATH):
    os.makedirs(OUTPUT_DIR_PATH)
    print(f"已建立資料夾: {OUTPUT_DIR_PATH}")

# 核心能力對應
competency_mapping = {
    'K1': ['A21', 'A22'],
    'K2': ['A23', 'A24'],
    'K3': ['A25', 'A26'],
    'K4': ['A27', 'A28'],
    'K5': ['A29', 'A210', 'A211']
}

core_competencies = {
    'K1': '能夠整合、組織電機專業理論來分析、表達問題之能力。',
    'K2': '能夠運用電機專業知識解決及實作電機工程問題之能力。',
    'K3': '具備分工、協調、重視團隊合作精神、遵守工程倫理以達成工作目標之能力。',
    'K4': '能夠激發自己潛能、融合他人智慧，具備獨立思考以及研究創新之能力。',
    'K5': '具備吸收電機新知、掌握國際發展趨勢，隨時接受競爭挑戰之能力。'
}

score_weights = {1: 20, 2: 40, 3: 60, 4: 80, 5: 100}

db_col_map = {
    'K1': 'has_SO_K1',
    'K2': 'has_SO_K2',
    'K3': 'has_SO_K3',
    'K4': 'has_SO_K4',
    'K5': 'has_SO_K5'
}

# ==========================================
# 2. 計算邏輯模組
# ==========================================

def calculate_survey_scores(df_source):
    """計算問卷分數 (依學年)"""
    if df_source.empty:
        return pd.DataFrame()

    df = df_source[~df_source['sem'].str.endswith('T')].copy()
    df['year'] = df['sem'].str[:3]
    
    yearly_stats = {}
    years = sorted(list(set(df['year'])))
    
    for year in years:
        yearly_stats[year] = {}
        for k_key in competency_mapping.keys():
            yearly_stats[year][k_key] = {'weighted_sum': 0, 'total_samples': 0}

    for k_key, qid_list in competency_mapping.items():
        for qid in qid_list:
            df_q = df[df['qid'] == qid]
            for _, row in df_q.iterrows():
                year = row['year']
                total_count = row['total']
                if total_count == 0: continue
                
                w_sum = 0
                for ans_idx in range(1, 6):
                    count = row[f'count_{ans_idx}']
                    w_sum += count * score_weights[ans_idx]
                
                yearly_stats[year][k_key]['weighted_sum'] += w_sum
                yearly_stats[year][k_key]['total_samples'] += total_count

    data = []
    for year in years:
        row = {'學年': year}
        for k_key in competency_mapping.keys():
            stats = yearly_stats[year][k_key]
            if stats['total_samples'] > 0:
                score = round(stats['weighted_sum'] / stats['total_samples'], 2)
            else:
                score = 0.0
            row[k_key] = score
        data.append(row)
        
    return pd.DataFrame(data)

def calculate_grade_scores(df_matrix):
    """計算成績分數 (依學年, 排除暑修)"""
    if df_matrix.empty:
        return pd.DataFrame()
        
    df = df_matrix[df_matrix['semester'] != 3].copy()
    df = df[df['course_score_AVG'].notnull()]
    
    years = sorted(df['academic_year'].unique())
    data = []
    
    for year in years:
        df_year = df[df['academic_year'] == year]
        row = {'學年': str(year)}
        
        for k_key in ['K1', 'K2', 'K3', 'K4', 'K5']:
            col_name = db_col_map[k_key]
            mask = df_year[col_name].apply(lambda x: bool(x) if pd.notnull(x) else False)
            df_k = df_year[mask]
            
            if not df_k.empty:
                avg = df_k['course_score_AVG'].mean()
                row[k_key] = round(avg, 2)
            else:
                row[k_key] = 0.0
        
        data.append(row)
    
    return pd.DataFrame(data)

# ==========================================
# 3. Excel 輸出模組
# ==========================================

def write_trend_table(ws, title, df_data, start_row, start_col=1):
    """
    [歷年趨勢表] 維持原本格式 (學年往下長)
    """
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+5)
    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    headers = ['學年'] + [f"{k} {desc}" for k, desc in core_competencies.items()]
    
    current_row = start_row + 1
    for i, h in enumerate(headers):
        cell = ws.cell(row=current_row, column=start_col + i, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        col_letter = chr(64 + start_col + i)
        if i == 0:
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = 25
            
    current_row += 1
    
    if df_data.empty:
        ws.cell(row=current_row, column=start_col, value="(無資料)")
        return current_row + 1

    for _, row in df_data.iterrows():
        cell = ws.cell(row=current_row, column=start_col, value=f"{row['學年']}學年度")
        cell.alignment = Alignment(horizontal='center')
        
        for i, k in enumerate(['K1', 'K2', 'K3', 'K4', 'K5'], 1):
            val = row.get(k, 0.0)
            cell = ws.cell(row=current_row, column=start_col + i, value=val)
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0.00'
            
        current_row += 1
        
    return current_row

def write_yearly_comparison_table(ws, year, row_survey, row_grade, start_row=1):
    """
    [單一學年比較表]
    第一欄：評量類型
    第一列：畢業生問券
    第二列：應屆畢業生成績
    """
    
    # 標題
    title = f"【{year}學年度 核心能力達成度綜合分析】"
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    
    current_row = start_row + 1
    
    # 欄位名稱
    headers = ['評量類型'] + [f"{k} {desc}" for k, desc in core_competencies.items()]
    
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=i, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        col_letter = chr(64 + i)
        if i == 1:
            ws.column_dimensions[col_letter].width = 25
        else:
            ws.column_dimensions[col_letter].width = 30
            
    current_row += 1
    
    # --- 第一列：畢業生問券 ---
    ws.cell(row=current_row, column=1, value="畢業生問券分析").alignment = Alignment(horizontal='center')
    
    if not row_survey.empty:
        # row_survey 是一個 DataFrame (即使只有一列)，取第一列
        data = row_survey.iloc[0]
        for i, k in enumerate(['K1', 'K2', 'K3', 'K4', 'K5'], 2):
            val = data.get(k, 0.0)
            cell = ws.cell(row=current_row, column=i, value=val)
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0.00'
    else:
        for i in range(2, 7):
            ws.cell(row=current_row, column=i, value="-").alignment = Alignment(horizontal='center')
            
    current_row += 1
    
    # --- 第二列：應屆畢業生成績 ---
    ws.cell(row=current_row, column=1, value="應屆畢業生成績分析").alignment = Alignment(horizontal='center')
    
    if not row_grade.empty:
        data = row_grade.iloc[0]
        for i, k in enumerate(['K1', 'K2', 'K3', 'K4', 'K5'], 2):
            val = data.get(k, 0.0)
            cell = ws.cell(row=current_row, column=i, value=val)
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0.00'
    else:
        for i in range(2, 7):
            ws.cell(row=current_row, column=i, value="-").alignment = Alignment(horizontal='center')
            
    return current_row

def export_integrated_report(df_survey, df_grade, filename_prefix):
    output_filename = f'{filename_prefix}_核心能力綜合分析_{today_str}.xlsx'
    full_path = os.path.join(OUTPUT_DIR_PATH, output_filename)
    
    print(f"正在輸出: {full_path}")
    
    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
        
        # --- 分頁 1: 歷年趨勢分析 (維持原樣，上下表方便畫圖) ---
        ws_trend = writer.book.create_sheet("歷年趨勢分析", 0)
        
        current_row = 1
        current_row = write_trend_table(ws_trend, f"【{filename_prefix} 畢業生問券分析 - 歷年變化】", df_survey, current_row)
        
        current_row += 2 
        current_row = write_trend_table(ws_trend, f"【{filename_prefix} 畢業生成績分析 - 歷年變化】", df_grade, current_row)
        
        # --- 分頁 2~N: 各學年比較表 (合併為一表) ---
        years_survey = set(df_survey['學年']) if not df_survey.empty else set()
        years_grade = set(df_grade['學年']) if not df_grade.empty else set()
        all_years = sorted(list(years_survey | years_grade))
        
        for year in all_years:
            sheet_name = f"{year}學年度"
            ws_year = writer.book.create_sheet(sheet_name)
            
            # 篩選資料
            row_survey = df_survey[df_survey['學年'] == year] if not df_survey.empty else pd.DataFrame()
            row_grade = df_grade[df_grade['學年'] == year] if not df_grade.empty else pd.DataFrame()
            
            # 寫入合併表格
            write_yearly_comparison_table(ws_year, year, row_survey, row_grade)
            
        # 移除預設 Sheet
        if 'Sheet' in writer.book.sheetnames:
            writer.book.remove(writer.book['Sheet'])

# ==========================================
# 4. 主程式執行
# ==========================================
print("讀取資料庫...")
df_u_survey = pd.read_sql("SELECT * FROM LDUdataAnalyze", db.conn)
df_g_survey = pd.read_sql("SELECT * FROM LDGdataAnalyze", db.conn)

sql_matrix = """
SELECT M.*, C.dept_code 
FROM Course_Matrix AS M
INNER JOIN Courses AS C ON M.course_id = C.id
"""
df_matrix_all = pd.read_sql(sql_matrix, db.conn)
if not df_matrix_all.empty:
    df_matrix_all['dept_code'] = df_matrix_all['dept_code'].astype(str).str.strip()

# --- 大學部 ---
print("\n--- 正在處理大學部資料 ---")
df_survey_u_res = calculate_survey_scores(df_u_survey)
df_grade_u_raw = df_matrix_all[df_matrix_all['dept_code'] == 'B301'].copy()
df_grade_u_res = calculate_grade_scores(df_grade_u_raw)
export_integrated_report(df_survey_u_res, df_grade_u_res, "大學部")

# --- 碩士班 ---
print("\n--- 正在處理碩士班資料 ---")
df_survey_g_res = calculate_survey_scores(df_g_survey)
df_grade_g_raw = df_matrix_all[df_matrix_all['dept_code'] == 'M301'].copy()
df_grade_g_res = calculate_grade_scores(df_grade_g_raw)
export_integrated_report(df_survey_g_res, df_grade_g_res, "碩士班")

db.close()
print("-" * 30)
print(f"作業完成！檔案已輸出至: {OUTPUT_DIR_PATH}")