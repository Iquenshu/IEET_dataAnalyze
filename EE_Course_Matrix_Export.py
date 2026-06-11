import pandas as pd
from Accessdb import AccessHelper
import os
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import warnings

# ==========================================
# 0. 全域路徑與目標設定
# ==========================================
# 目標統計的學年度列表
TARGET_YEARS = [str(y) for y in range(109, 115)]  # 109 ~ 114 學年度

# 輸出目錄與檔名設定
OUTPUT_BASE_DIR = 'output_files'
OUTPUT_SUB_DIR = '畢業生成績分析'

# ==========================================
# 1. 環境與環境初始化
# ==========================================
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)

db = AccessHelper()
today_str = datetime.today().strftime('%Y%m%d')

# 自動建立輸出目錄
OUTPUT_DIR_PATH = os.path.join(OUTPUT_BASE_DIR, OUTPUT_SUB_DIR)
if not os.path.exists(OUTPUT_DIR_PATH):
    os.makedirs(OUTPUT_DIR_PATH)
    print(f"建立目錄: {OUTPUT_DIR_PATH}")

# ==========================================
# 2. 核心匯出清冊邏輯
# ==========================================
def export_ee_course_matrix_standalone(target_years, filename):
    full_path = os.path.join(OUTPUT_DIR_PATH, filename)
    print(f"正在從資料庫撈取 Courses 歷年完整開課資料...")
    
    # 讀取 Access 資料庫中的 Courses 完整實體欄位
    sql_courses = """
        SELECT academic_year, semester, course_code, course_name, is_required, credits, instructor, 
               is_math, is_science, is_eng_prof, is_general 
        FROM Courses
    """
    df_courses = pd.read_sql(sql_courses, db.conn)
    
    if df_courses.empty:
        print("⚠️ 警告：資料庫 Courses 資料表內無任何數據，無法匯出清冊。")
        return
        
    print(f"正在建立與匯出電機系開課分類清冊: {full_path}")
    
    # A. 篩選目標學年度內的電機系開課紀錄
    df_filtered = df_courses[df_courses['academic_year'].astype(str).isin(target_years)].copy()
    
    # 依 學年度、學期、課號 進行整齊排序
    df_filtered = df_filtered.sort_values(by=['academic_year', 'semester', 'course_code'])
    
    # B. 定義與重命名欲產出的報表欄位
    cols_mapping = {
        'academic_year': '學年度',
        'semester': '學期',
        'course_code': '課號',
        'course_name': '課程名稱',
        'is_required': '必選修',
        'credits': '學分',
        'instructor': '授課教師',
        'is_math': '數學',
        'is_science': '基礎科學',
        'is_eng_prof': '工程專業',
        'is_general': '通識'
    }
    
    df_export = df_filtered[list(cols_mapping.keys())].rename(columns=cols_mapping)
    
    # C. 將分類的布林值 True/False 轉換為 1 與 0 數字 (方便後續統計與權重乘法)
    for cat in ['數學', '基礎科學', '工程專業', '通識']:
        df_export[cat] = df_export[cat].apply(lambda x: 1 if x else 0)
        
    # D. 利用 openpyxl 進行美化排版輸出
    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='電機系開課分類矩陣')
        ws = writer.sheets['電機系開課分類矩陣']
        
        # 設定經典質感淺藍色標頭與細邊框
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 標頭樣式調整
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        # 內文樣式、對齊、與框線設定
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                # 置中控制：除了課程名稱外，其餘欄位一律全面置中
                if cell.column in [1, 2, 3, 5, 6, 8, 9, 10, 11]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
        # 自動適應欄寬度計算
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 11)
            
    print(f"🎉 歷年開課課程分類清冊導出成功！")

# ==========================================
# 3. 主執行程序入口
# ==========================================
if __name__ == "__main__":
    out_filename = f"電機系開課課程分類清冊_{today_str}.xlsx"
    export_ee_course_matrix_standalone(TARGET_YEARS, out_filename)
    db.close()
    print("程序關閉，完成。")